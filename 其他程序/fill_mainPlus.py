import sys
import os
import pandas as pd
import numpy as np
import joblib
import configparser
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext, Listbox, MULTIPLE, SINGLE, END
import json
from datetime import datetime
import re
import string
from sklearn.feature_extraction.text import TfidfVectorizer, CountVectorizer
from sklearn.preprocessing import OneHotEncoder
from sklearn.ensemble import RandomForestClassifier
from sklearn.multioutput import MultiOutputClassifier
import traceback
from openpyxl.styles import Font
import codecs
from collections import defaultdict
import hashlib
import warnings
warnings.filterwarnings('ignore')

# ==================== 全局配置 ====================
def get_real_base_dir():
    """获取真正的根目录，兼容各种运行模式"""
    if hasattr(sys, 'frozen'):
        original_exe_path = os.path.abspath(sys.executable)
        original_exe_dir = os.path.dirname(original_exe_path)
    else:
        script_path = os.path.abspath(__file__)
        original_exe_dir = os.path.dirname(script_path)
    
    base_dir = os.path.abspath(os.path.join(original_exe_dir, ".."))
    return base_dir, original_exe_dir

BASE_DIR, ORIGINAL_FILE_DIR = get_real_base_dir()
DATA_DIR = os.path.join(BASE_DIR, "data")
CODE_DIR = os.path.join(BASE_DIR, "code")
MODEL_DIR = os.path.join(BASE_DIR, "model")
MAPPING_DIR = os.path.join(BASE_DIR, "mapping_rules")
LOG_DIR = os.path.join(BASE_DIR, "logs")
INCREMENTAL_DIR = os.path.join(DATA_DIR, "incremental_data")

# 文件路径
HISTORY_EXCEL = os.path.join(DATA_DIR, "history_data.xlsx")
NEW_EXCEL = os.path.join(DATA_DIR, "new_data.xlsx")
CONFIG_PATH = os.path.join(CODE_DIR, "config.ini")
LOG_JSON_PATH = os.path.join(LOG_DIR, "operation_log.json")
LOG_TXT_PATH = os.path.join(LOG_DIR, "operation_log.txt")
LOW_CONF_OUTPUT = os.path.join(DATA_DIR, "低置信数据校核清单.xlsx")
INCREMENTAL_OUTPUT = os.path.join(INCREMENTAL_DIR, "增量数据_待复核.xlsx")
INCREMENTAL_CONFIG = os.path.join(INCREMENTAL_DIR, "增量配置.json")

# 置信度阈值
CONFIDENCE_THRESHOLD = 0.7

# 自动创建必要文件夹
for folder in [DATA_DIR, CODE_DIR, MODEL_DIR, MAPPING_DIR, LOG_DIR, INCREMENTAL_DIR]:
    if not os.path.exists(folder):
        os.makedirs(folder)

# ==================== 核心算法函数 ====================

def preprocess_text(text_series):
    """纯函数文本预处理"""
    # 填充空值+转字符串
    text_series = text_series.fillna("未知").astype(str)
    # 移除标点/数字/多余空格
    text_series = text_series.apply(lambda x: re.sub(f"[{re.escape(string.punctuation)}]", " ", x))
    text_series = text_series.apply(lambda x: re.sub(r"\d+", "", x))
    text_series = text_series.apply(lambda x: re.sub(r"\s+", " ", x).strip().lower())
    return text_series

def extract_text_features(text_series, tfidf_params, fit=True, tfidf_model=None, count_model=None):
    """纯函数提取文本特征"""
    # 预处理文本
    text_clean = preprocess_text(text_series)
    
    # TF-IDF特征
    if fit:
        tfidf = TfidfVectorizer(
            ngram_range=tfidf_params["ngram_range"],
            max_features=tfidf_params["max_features"],
            stop_words=tfidf_params["stop_words"]
        )
        tfidf_feat = tfidf.fit_transform(text_clean).toarray()
    else:
        tfidf_feat = tfidf_model.transform(text_clean).toarray()
    
    # Count特征
    if fit:
        count = CountVectorizer(ngram_range=(1, 1), max_features=50)
        count_feat = count.fit_transform(text_clean).toarray()
    else:
        count_feat = count_model.transform(text_clean).toarray()
    
    # 统计特征
    stats = pd.DataFrame()
    stats['length'] = text_clean.apply(len)
    stats['word_count'] = text_clean.apply(lambda x: len(set(x.split())) if x else 0)
    stats['unique_ratio'] = text_clean.apply(lambda x: len(set(x.split()))/len(x.split()) if x else 0)
    stats_feat = stats.values
    
    # 拼接特征
    text_feat = np.hstack([tfidf_feat, count_feat, stats_feat])
    
    if fit:
        return text_feat, tfidf, count
    else:
        return text_feat

def clean_excel_data(df, is_train_data=False, target_cols=None, text_columns=None):
    """增强版数据清洗：保护身份证号等长数字文本，避免科学计数法"""
    print(f"🧹 清洗数据（原始行数：{len(df)}）")
    
    # 1. 首先备份原始数据类型
    original_dtypes = df.dtypes.to_dict()
    
    # 2. 识别可能需要保护的列（身份证、手机号等长数字文本）
    protected_columns = []
    if text_columns:
        for col in text_columns:
            if col in df.columns:
                # 检查列名是否包含身份证、手机号等关键词
                col_lower = col.lower()
                if any(keyword in col_lower for keyword in ['身份证', '身份证号', '身份证号码', 'id', 'card', '手机', '电话', 'phone', 'tel']):
                    protected_columns.append(col)
    
    # 3. 填充空值，但对于保护列，保持为字符串
    for col in df.columns:
        if df[col].dtype == 'object' or col in protected_columns:
            # 对于对象类型或保护列，填充为字符串"未知"
            df[col] = df[col].fillna("未知").replace("", "未知")
            
            # 对于保护列，确保是字符串格式，避免科学计数法
            if col in protected_columns:
                df[col] = df[col].astype(str).apply(lambda x: x.strip())
                # 处理长数字文本：如果是纯数字且长度大于10，转换为字符串并去掉科学计数法
                df[col] = df[col].apply(lambda x: 
                    str(int(float(x))) if (re.match(r'^\d+\.0$', str(x)) or re.match(r'^\d+\.\d+$', str(x))) and len(str(int(float(x)))) >= 15
                    else str(x))
        else:
            # 对于数值类型，填充为0
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    # 4. 训练数据清洗目标列
    if is_train_data and target_cols:
        for col in target_cols:
            if col in df.columns:
                df[col] = df[col].fillna("未知").replace("", "未知")
                # 如果目标列在保护列表中，确保是字符串
                if col in protected_columns:
                    df[col] = df[col].astype(str)
    
    # 5. 删除全空行
    df = df.dropna(how='all').reset_index(drop=True)
    
    # 6. 去除重复行，但保留至少一行
    original_len = len(df)
    df = df.drop_duplicates()
    removed_duplicates = original_len - len(df)
    if removed_duplicates > 0:
        print(f"⚠️  移除了 {removed_duplicates} 条重复记录")
    
    print(f"✅ 清洗完成（剩余行数：{len(df)}）")
    return df

def build_mapping_dict(source_file_path, key_cols, value_col, clean_data=True):
    """从源数据构建多列→单列的精确映射字典"""
    # 读取源数据 - 修复：指定引擎
    if source_file_path.endswith('.csv'):
        df = pd.read_csv(source_file_path, dtype=str)  # 强制所有列读取为字符串
    else:
        # 明确指定引擎，先尝试openpyxl，再尝试xlrd
        try:
            df = pd.read_excel(source_file_path, dtype=str, engine='openpyxl')
        except Exception:
            try:
                df = pd.read_excel(source_file_path, dtype=str, engine='xlrd')
            except Exception as e:
                raise ValueError(f"无法读取Excel文件：{source_file_path}，错误：{str(e)}")
    
    # 数据清洗
    if clean_data:
        df = clean_excel_data(df, is_train_data=False)
    
    # 验证列是否存在
    missing_cols = [col for col in key_cols + [value_col] if col not in df.columns]
    if missing_cols:
        raise ValueError(f"源数据中缺少列：{', '.join(missing_cols)}")
    
    # 去除空值+去重
    df = df.dropna(subset=key_cols + [value_col])
    df = df.drop_duplicates(subset=key_cols, keep='first')
    
    # 构建复合键（元组形式）
    if len(key_cols) == 1:
        mapping_dict = dict(zip(df[key_cols[0]], df[value_col]))
        print(f"✅ 构建映射字典完成：共 {len(mapping_dict)} 条唯一映射，键列：{key_cols}，值列：{value_col}")
        return mapping_dict
    else:
        df['composite_key'] = df[key_cols].apply(tuple, axis=1)
        # 构建映射字典
        mapping_dict = dict(zip(df['composite_key'], df[value_col]))
        print(f"✅ 构建映射字典完成：共 {len(mapping_dict)} 条唯一映射，键列：{key_cols}，值列：{value_col}")
        return mapping_dict

def lookup_value(df, mapping_dict, key_cols, value_col, fill_na="未知"):
    """对新数据进行精确查表填充"""
    # 验证键列是否存在
    missing_cols = [col for col in key_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"新数据中缺少键列：{', '.join(missing_cols)}")
    
    # 核心优化：若目标列不存在，先在DataFrame中创建该列
    if value_col not in df.columns:
        df[value_col] = fill_na
    
    # 构建复合键
    if len(key_cols) == 1:
        df[value_col] = df[key_cols[0]].map(mapping_dict)
    else:
        df['composite_key'] = df[key_cols].apply(tuple, axis=1)
        # 精确查表填充
        df[value_col] = df['composite_key'].map(mapping_dict)
        # 清理辅助列
        df = df.drop(columns=['composite_key'])
    
    df[value_col] = df[value_col].fillna(fill_na)
    
    return df

def train_model(feat_cols, text_cols, target_cols, model_params, history_excel_path):
    """纯基础模型训练"""
    if not target_cols:
        raise ValueError("模型训练失败：必须选择至少1个目标列！")
    all_features = feat_cols + text_cols
    if not all_features:
        raise ValueError("模型训练失败：无有效特征列，无法提取特征！")
    
    print(f"\n🚀 开始训练模型...")
    
    # 读取并清洗训练数据 - 修复：指定引擎
    try:
        train_df = pd.read_excel(history_excel_path, dtype=str, engine='openpyxl')  # 所有列读取为字符串
    except Exception:
        try:
            train_df = pd.read_excel(history_excel_path, dtype=str, engine='xlrd')
        except Exception as e:
            raise ValueError(f"无法读取训练数据文件：{history_excel_path}，错误：{str(e)}")
    
    train_df = clean_excel_data(train_df, is_train_data=True, target_cols=target_cols, text_columns=text_cols)
    
    # 1. 处理普通特征（OneHot编码）
    X_cat = None
    encoder = None
    if feat_cols:
        encoder = OneHotEncoder(handle_unknown="ignore", sparse_output=False)
        X_cat = encoder.fit_transform(train_df[feat_cols])
    
    # 2. 处理文本特征
    X_text_all = None
    tfidf_dict = {}
    count_dict = {}
    if text_cols:
        tfidf_params = model_params["tfidf"]
        for text_col in text_cols:
            text_feat, tfidf, count = extract_text_features(train_df[text_col], tfidf_params)
            tfidf_dict[text_col] = tfidf
            count_dict[text_col] = count
            if X_text_all is None:
                X_text_all = text_feat
            else:
                X_text_all = np.hstack([X_text_all, text_feat])
    
    # 3. 拼接所有特征
    X_list = []
    if X_cat is not None:
        X_list.append(X_cat)
    if X_text_all is not None:
        X_list.append(X_text_all)
    if not X_list:
        raise ValueError("无有效特征数据，无法训练模型！")
    X_train = np.hstack(X_list)
    y_train = train_df[target_cols]
    
    # 4. 训练分类器
    rf_params = model_params["rf"]
    base_clf = RandomForestClassifier(
        n_estimators=rf_params["n_estimators"],
        max_depth=rf_params["max_depth"],
        min_samples_split=rf_params["min_samples_split"],
        min_samples_leaf=rf_params["min_samples_leaf"],
        random_state=rf_params["random_state"],
        n_jobs=rf_params["n_jobs"]
    )
    if len(target_cols) > 1:
        classifier = MultiOutputClassifier(base_clf)
    else:
        classifier = base_clf
    classifier.fit(X_train, y_train)
    
    print(f"✅ 模型训练完成（训练数据：{len(train_df)}行，特征维度：{X_train.shape[1]}）")
    return encoder, tfidf_dict, count_dict, classifier, len(train_df)

def predict_data(encoder, tfidf_dict, count_dict, classifier, feat_cols, text_cols, 
                 target_cols, mapping_config, model_params, new_excel_path, progress_callback=None):
    """
    执行模型预测（核心优化：先查表填充，再模型预测）
    """
    # 步骤1：读取并清洗新数据（所有列读取为字符串） - 修复：指定引擎
    try:
        new_df = pd.read_excel(new_excel_path, dtype=str, engine='openpyxl')
    except Exception:
        try:
            new_df = pd.read_excel(new_excel_path, dtype=str, engine='xlrd')
        except Exception as e:
            raise ValueError(f"无法读取新数据文件：{new_excel_path}，错误：{str(e)}")
    
    new_df = clean_excel_data(new_df, text_columns=text_cols)
    if len(new_df) == 0:
        raise ValueError("无有效预测数据，无法进行处理！")
    
    if progress_callback:
        progress_callback(10, "读取和清洗数据完成")
    
    # 步骤2：先执行「查表填充」
    mapped_target_cols = []
    if mapping_config and len(mapping_config) > 0:
        if progress_callback:
            progress_callback(20, "执行查表填充...")
        
        for idx, (source_file, key_cols, value_col) in enumerate(mapping_config):
            try:
                mapping_dict = build_mapping_dict(source_file, key_cols, value_col)
                new_df = lookup_value(new_df, mapping_dict, key_cols, value_col)
                mapped_target_cols.append(value_col)
                
                if progress_callback:
                    progress = 20 + (idx + 1) * 30 / len(mapping_config)
                    progress_callback(progress, f"查表填充 {idx+1}/{len(mapping_config)}: {value_col}")
                    
            except Exception as e:
                print(f"⚠️  查表填充失败 [{key_cols}->{value_col}]: {str(e)}")
        
        print(f"✅ 查表填充完成，共生成/填充 {len(mapped_target_cols)} 个目标列")
    
    if progress_callback:
        progress_callback(50, "查表填充完成")
    
    # 步骤3：后执行「模型预测」（跳过已查表填充的列）
    model_target_cols = [col for col in target_cols if col not in mapped_target_cols]
    all_features = feat_cols + text_cols
    
    if model_target_cols and all_features:
        # 3.1 提取模型特征
        if progress_callback:
            progress_callback(60, "提取模型特征...")
        
        X_cat = None
        if feat_cols:
            valid_feat_cols = [col for col in feat_cols if col in new_df.columns]
            if valid_feat_cols and encoder:
                X_cat = encoder.transform(new_df[valid_feat_cols])
        
        # 3.2 处理文本特征
        X_text_all = None
        tfidf_params = model_params["tfidf"]
        if text_cols:
            valid_text_cols = [col for col in text_cols if col in new_df.columns]
            for text_col in valid_text_cols:
                if text_col in tfidf_dict and text_col in count_dict:
                    text_feat = extract_text_features(new_df[text_col], tfidf_params, 
                                                     fit=False, tfidf_model=tfidf_dict[text_col], 
                                                     count_model=count_dict[text_col])
                    if X_text_all is None:
                        X_text_all = text_feat
                    else:
                        X_text_all = np.hstack([X_text_all, text_feat])
        
        # 3.3 拼接特征
        if progress_callback:
            progress_callback(70, "拼接预测特征...")
        
        X_list = []
        if X_cat is not None:
            X_list.append(X_cat)
        if X_text_all is not None:
            X_list.append(X_text_all)
        if not X_list:
            raise ValueError("无有效特征数据，无法进行模型预测！")
        X_new = np.hstack(X_list)
        
        # 3.4 模型预测
        if progress_callback:
            progress_callback(80, "执行模型预测...")
        
        pred_results = classifier.predict(X_new)
        pred_proba = classifier.predict_proba(X_new)
        
        # 3.5 计算置信度并组装结果
        if progress_callback:
            progress_callback(90, "计算置信度...")
        
        confidence = []
        if len(model_target_cols) == 1:
            confidence = [round(max(p), 4) for p in pred_proba]
        else:
            for i in range(len(pred_results)):
                avg_proba = np.mean([max(proba[i]) for proba in pred_proba])
                confidence.append(round(avg_proba, 4))
        
        new_df["置信度"] = confidence
        if len(model_target_cols) == 1:
            new_df[model_target_cols[0]] = pred_results
        else:
            for idx, col in enumerate(model_target_cols):
                new_df[col] = pred_results[:, idx]
        
        print(f"✅ 模型预测完成，生成置信度列")
        
    elif model_target_cols and not all_features:
        raise ValueError("模型预测失败：存在模型目标列，但未选择任何有效特征列！")
    else:
        if progress_callback:
            progress_callback(90, "所有目标列均为查表填充，跳过模型预测...")
        # 如果没有模型预测，添加一个默认的置信度列
        new_df["置信度"] = 1.0
        print("⚠️  所有目标列均为查表填充，置信度设为1.0")
    
    if progress_callback:
        progress_callback(95, "处理完成，正在保存结果...")
    
    return new_df, len(new_df)

def write_results_to_excel(df, output_path):
    """将结果写入Excel文件，保护长数字文本格式 - 修复版本"""
    try:
        # 创建写入器 - 使用新版本的API
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 将所有列转换为字符串，避免科学计数法
            for col in df.columns:
                if df[col].dtype == 'object':
                    df[col] = df[col].astype(str)
            
            # 写入数据
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # 获取工作表
            worksheet = writer.book['Sheet1'] if 'Sheet1' in writer.book.sheetnames else writer.book.active
            
            # 标红低置信度记录
            if "置信度" in df.columns:
                conf_col = df.columns.get_loc("置信度") + 1
                for row in range(2, len(df)+2):
                    cell = worksheet.cell(row=row, column=conf_col)
                    try:
                        if cell.value and float(cell.value) < CONFIDENCE_THRESHOLD:
                            cell.font = Font(color="FF0000", bold=True)
                    except:
                        pass
            
            # 保护长数字文本列（身份证、手机号等）
            for col_idx, col_name in enumerate(df.columns, start=1):
                col_lower = col_name.lower()
                if any(keyword in col_lower for keyword in ['身份证', '身份证号', '身份证号码', 'id', 'card', '手机', '电话', 'phone', 'tel']):
                    # 设置这些列为文本格式
                    for row in range(2, len(df)+2):
                        cell = worksheet.cell(row=row, column=col_idx)
                        cell.number_format = '@'  # 文本格式
            
            # 保存文件
            writer.save()
        
        print(f"✅ 结果已保存到：{output_path}")
        
    except Exception as e:
        # 如果上述方法失败，使用更简单的方法
        print(f"⚠️  使用简化方法保存Excel：{str(e)}")
        try:
            # 直接使用to_excel方法
            df.to_excel(output_path, index=False, engine='openpyxl')
            print(f"✅ 结果已保存到：{output_path}")
        except Exception as e2:
            print(f"❌ 保存失败：{str(e2)}")
            raise e2

# ==================== 配置管理函数 ====================
def load_config():
    """加载配置文件"""
    config = configparser.ConfigParser()
    if os.path.exists(CONFIG_PATH):
        config.read(CONFIG_PATH, encoding="utf-8")
    feat_cols = config["FEATURES"]["cols"].split(",") if "FEATURES" in config else []
    target_cols = config["TARGETS"]["cols"].split(",") if "TARGETS" in config else []
    text_cols = config["TEXT_FEATURES"]["cols"].split(",") if "TEXT_FEATURES" in config else []
    # 处理空字符串
    feat_cols = [col for col in feat_cols if col.strip()]
    target_cols = [col for col in target_cols if col.strip()]
    text_cols = [col for col in text_cols if col.strip()]
    return feat_cols, target_cols, text_cols

def save_config(feat_cols, target_cols, text_cols):
    """保存配置文件"""
    config = configparser.ConfigParser()
    config["FEATURES"] = {"cols": ",".join(feat_cols)}
    config["TARGETS"] = {"cols": ",".join(target_cols)}
    config["TEXT_FEATURES"] = {"cols": ",".join(text_cols)}
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        config.write(f)

# ==================== 新增：数据对比器类 ====================
class DataComparator:
    """数据对比器：对比新数据与历史数据，识别增量部分"""
    
    def __init__(self, history_data_path, new_data_path):
        self.history_data_path = history_data_path
        self.new_data_path = new_data_path
        self.comparison_results = {}
        
    def compare_features(self, feature_cols, text_cols):
        """对比特征列的差异"""
        try:
            # 读取数据（所有列作为字符串读取） - 修复：指定引擎
            try:
                history_df = pd.read_excel(self.history_data_path, dtype=str, engine='openpyxl')
                new_df = pd.read_excel(self.new_data_path, dtype=str, engine='openpyxl')
            except Exception:
                try:
                    history_df = pd.read_excel(self.history_data_path, dtype=str, engine='xlrd')
                    new_df = pd.read_excel(self.new_data_path, dtype=str, engine='xlrd')
                except Exception as e:
                    raise ValueError(f"无法读取数据文件：{str(e)}")
            
            comparison = {
                "feature_differences": {},
                "summary": {
                    "history_rows": len(history_df),
                    "new_rows": len(new_df),
                    "total_differences": 0
                }
            }
            
            # 对比所有特征列
            all_cols = list(set(feature_cols + text_cols))
            
            for col in all_cols:
                if col in history_df.columns and col in new_df.columns:
                    hist_vals = set(history_df[col].dropna().astype(str).str.strip().unique())
                    new_vals = set(new_df[col].dropna().astype(str).str.strip().unique())
                    
                    # 找出新数据中独有的值
                    new_unique_vals = new_vals - hist_vals
                    
                    if new_unique_vals:
                        comparison["feature_differences"][col] = {
                            "new_unique_count": len(new_unique_vals),
                            "new_unique_values": list(new_unique_vals)[:10],  # 只显示前10个
                            "history_unique_count": len(hist_vals),
                            "new_data_unique_count": len(new_vals)
                        }
                        comparison["summary"]["total_differences"] += 1
            
            self.comparison_results["features"] = comparison
            return comparison
            
        except Exception as e:
            print(f"特征对比失败：{str(e)}")
            return None
    
    def compare_results(self, target_cols, result_df):
        """对比预测结果与历史数据的差异"""
        try:
            # 修复：指定引擎
            try:
                history_df = pd.read_excel(self.history_data_path, dtype=str, engine='openpyxl')
            except Exception:
                try:
                    history_df = pd.read_excel(self.history_data_path, dtype=str, engine='xlrd')
                except Exception as e:
                    raise ValueError(f"无法读取历史数据文件：{str(e)}")
            
            comparison = {
                "result_differences": {},
                "summary": {
                    "history_targets_count": {},
                    "new_predictions_count": {},
                    "new_categories": {}
                }
            }
            
            for target_col in target_cols:
                if target_col in history_df.columns and target_col in result_df.columns:
                    # 历史数据中的类别分布
                    hist_categories = history_df[target_col].value_counts().to_dict()
                    
                    # 新数据中的预测类别分布
                    new_categories = result_df[target_col].value_counts().to_dict()
                    
                    # 找出新预测中独有的类别
                    hist_cat_set = set(hist_categories.keys())
                    new_cat_set = set(new_categories.keys())
                    new_unique_cats = new_cat_set - hist_cat_set
                    
                    comparison["result_differences"][target_col] = {
                        "history_categories": hist_categories,
                        "new_categories": new_categories,
                        "new_unique_categories": list(new_unique_cats) if new_unique_cats else [],
                        "new_unique_count": len(new_unique_cats)
                    }
                    
                    comparison["summary"]["history_targets_count"][target_col] = sum(hist_categories.values())
                    comparison["summary"]["new_predictions_count"][target_col] = sum(new_categories.values())
                    comparison["summary"]["new_categories"][target_col] = len(new_unique_cats)
            
            self.comparison_results["results"] = comparison
            return comparison
            
        except Exception as e:
            print(f"结果对比失败：{str(e)}")
            return None
    
    def extract_incremental_data(self, result_df, feature_cols, target_cols, confidence_threshold=CONFIDENCE_THRESHOLD):
        """提取需要人工复核的增量数据"""
        try:
            # 修复：指定引擎
            try:
                history_df = pd.read_excel(self.history_data_path, dtype=str, engine='openpyxl')
            except Exception:
                try:
                    history_df = pd.read_excel(self.history_data_path, dtype=str, engine='xlrd')
                except Exception as e:
                    raise ValueError(f"无法读取历史数据文件：{str(e)}")
            
            incremental_data = []
            
            for idx, new_row in result_df.iterrows():
                is_incremental = False
                reasons = []
                
                # 1. 检查低置信度
                if "置信度" in new_row and new_row["置信度"] and float(new_row["置信度"]) < confidence_threshold:
                    is_incremental = True
                    reasons.append("低置信度")
                
                # 2. 检查特征值是否在历史数据中出现过
                for col in feature_cols:
                    if col in history_df.columns and col in new_row:
                        hist_vals = set(history_df[col].dropna().astype(str).str.strip().unique())
                        new_val = str(new_row[col]).strip()
                        if new_val not in hist_vals and new_val not in ["", "未知", "nan"]:
                            is_incremental = True
                            reasons.append(f"新特征值[{col}]")
                            break  # 找到一个新特征值就足够
                
                # 3. 检查预测结果是否在历史数据中出现过
                for target_col in target_cols:
                    if target_col in history_df.columns and target_col in new_row:
                        hist_categories = set(history_df[target_col].dropna().astype(str).str.strip().unique())
                        new_category = str(new_row[target_col]).strip()
                        if new_category not in hist_categories and new_category not in ["", "未知", "nan"]:
                            is_incremental = True
                            reasons.append(f"新预测类别[{target_col}]")
                            break
                
                if is_incremental:
                    # 复制行数据并添加原因
                    row_data = new_row.to_dict()
                    row_data["_复核原因"] = "、".join(reasons)
                    row_data["_数据来源"] = "新数据预测"
                    row_data["_时间戳"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    incremental_data.append(row_data)
            
            incremental_df = pd.DataFrame(incremental_data)
            
            if not incremental_df.empty:
                # 确保列的顺序
                cols_order = ["_复核原因", "_数据来源", "_时间戳"] + \
                            [col for col in result_df.columns if col not in ["_复核原因", "_数据来源", "_时间戳"]]
                incremental_df = incremental_df.reindex(columns=cols_order)
            
            return incremental_df
            
        except Exception as e:
            print(f"提取增量数据失败：{str(e)}")
            return pd.DataFrame()

# ==================== 新增：增量学习管理器类 ====================
class IncrementalLearner:
    """增量学习管理器：管理增量数据的收集、复核和模型更新"""
    
    def __init__(self, data_dir=INCREMENTAL_DIR):
        self.data_dir = data_dir
        self.pending_file = os.path.join(data_dir, "增量数据_待复核.xlsx")
        self.reviewed_file = os.path.join(data_dir, "增量数据_已复核.xlsx")
        self.config_file = os.path.join(data_dir, "增量配置.json")
        
        # 加载配置
        self.config = self.load_config()
    
    def load_config(self):
        """加载增量学习配置"""
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                pass
        
        # 默认配置
        return {
            "auto_incremental_learning": True,
            "confidence_threshold": CONFIDENCE_THRESHOLD,
            "min_incremental_samples": 10,
            "max_review_days": 7,
            "last_review_date": None,
            "total_reviewed": 0,
            "total_added": 0,
            "last_extraction_time": None,
            "preserve_edited_data": True,  # 新增：保护已编辑数据
            "last_append_date": None,  # 新增：最后追加日期
            "appended_records": []  # 新增：已追加记录ID，避免重复追加
        }
    
    def save_config(self):
        """保存增量学习配置"""
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"保存配置失败：{str(e)}")
            return False
    
    def save_incremental_data(self, incremental_df, description="自动提取", preserve_edited=True):
        """保存增量数据到待复核文件"""
        try:
            if incremental_df.empty:
                return False, "无增量数据需要复核"
            
            # 如果文件已存在且需要保护已编辑数据
            if os.path.exists(self.pending_file) and preserve_edited and self.config.get("preserve_edited_data", True):
                # 修复：指定引擎
                try:
                    existing_df = pd.read_excel(self.pending_file, dtype=str, engine='openpyxl')
                except Exception:
                    try:
                        existing_df = pd.read_excel(self.pending_file, dtype=str, engine='xlrd')
                    except Exception as e:
                        return False, f"无法读取待复核文件：{str(e)}"
                
                # 标记哪些行是已经编辑过的（有_复核状态标记的）
                if "_复核状态" in existing_df.columns:
                    # 已编辑的行（有复核状态）
                    edited_rows = existing_df[existing_df["_复核状态"].notna() & (existing_df["_复核状态"] != "")]
                    
                    # 未编辑的行
                    unedited_rows = existing_df[existing_df["_复核状态"].isna() | (existing_df["_复核状态"] == "")]
                    
                    # 只对未编辑的行进行去重合并
                    if not unedited_rows.empty:
                        # 使用特征列进行合并和去重
                        feature_columns = [col for col in incremental_df.columns if not col.startswith('_')]
                        
                        # 合并新数据和未编辑的旧数据
                        combined_undedited = pd.concat([unedited_rows[feature_columns], incremental_df[feature_columns]], ignore_index=True)
                        combined_undedited = combined_undedited.drop_duplicates(subset=feature_columns, keep='last')
                        
                        # 重新添加系统列
                        for sys_col in ["_复核原因", "_数据来源", "_时间戳", "_复核状态", "_复核时间", "_复核人员"]:
                            if sys_col in unedited_rows.columns:
                                combined_undedited[sys_col] = ""
                        
                        # 如果新数据有系统列，填充
                        for sys_col in incremental_df.columns:
                            if sys_col.startswith('_') and sys_col in incremental_df.columns:
                                combined_undedited[sys_col] = incremental_df[sys_col]
                        
                        # 合并已编辑的行和新的未编辑数据
                        combined_df = pd.concat([edited_rows, combined_undedited], ignore_index=True)
                    else:
                        # 所有行都已编辑，只添加全新的数据
                        feature_columns = [col for col in incremental_df.columns if not col.startswith('_')]
                        existing_features = edited_rows[feature_columns].apply(tuple, axis=1).tolist()
                        
                        # 只添加在已编辑数据中不存在的新数据
                        new_rows = []
                        for idx, row in incremental_df.iterrows():
                            row_tuple = tuple(row[col] for col in feature_columns)
                            if row_tuple not in existing_features:
                                new_rows.append(row)
                        
                        if new_rows:
                            new_df = pd.DataFrame(new_rows)
                            combined_df = pd.concat([edited_rows, new_df], ignore_index=True)
                        else:
                            combined_df = edited_rows.copy()
                else:
                    # 没有_复核状态列，直接去重合并
                    feature_columns = [col for col in incremental_df.columns if not col.startswith('_')]
                    combined_df = pd.concat([existing_df, incremental_df], ignore_index=True)
                    combined_df = combined_df.drop_duplicates(subset=feature_columns, keep='last')
            else:
                # 文件不存在或不需要保护已编辑数据
                combined_df = incremental_df.copy()
            
            # 保存到Excel（保护文本格式）
            write_results_to_excel(combined_df, self.pending_file)
            
            # 添加统计信息
            stats = {
                "提取时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "提取描述": description,
                "新增记录数": len(incremental_df),
                "总待复核记录数": len(combined_df),
                "特征列": list(incremental_df.columns),
                "保护已编辑数据": preserve_edited
            }
            
            # 更新配置
            self.config["last_extraction_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.save_config()
            
            print(f"✅ 增量数据已保存：{len(incremental_df)} 条记录待复核（保护已编辑数据：{preserve_edited}）")
            return True, stats
            
        except Exception as e:
            return False, f"保存增量数据失败：{str(e)}"
    
    def load_pending_data(self):
        """加载待复核数据（所有列作为字符串读取）"""
        try:
            if os.path.exists(self.pending_file):
                # 修复：指定引擎
                try:
                    return pd.read_excel(self.pending_file, dtype=str, engine='openpyxl')
                except Exception:
                    try:
                        return pd.read_excel(self.pending_file, dtype=str, engine='xlrd')
                    except Exception:
                        return pd.read_excel(self.pending_file, dtype=str)
            return pd.DataFrame()
        except:
            return pd.DataFrame()
    
    def load_reviewed_data(self):
        """加载已复核数据（所有列作为字符串读取）"""
        try:
            if os.path.exists(self.reviewed_file):
                # 修复：指定引擎
                try:
                    return pd.read_excel(self.reviewed_file, dtype=str, engine='openpyxl')
                except Exception:
                    try:
                        return pd.read_excel(self.reviewed_file, dtype=str, engine='xlrd')
                    except Exception:
                        return pd.read_excel(self.reviewed_file, dtype=str)
            return pd.DataFrame()
        except:
            return pd.DataFrame()
    
    def mark_as_reviewed(self, indices, correct_data=None):
        """标记数据为已复核"""
        try:
            pending_df = self.load_pending_data()
            if pending_df.empty:
                return False, "无待复核数据"
            
            # 检查索引是否有效
            valid_indices = [idx for idx in indices if idx in pending_df.index]
            if not valid_indices:
                return False, "无效的索引或所选数据已标记"
            
            # 检查是否已标记（通过_复核状态列判断）
            already_reviewed = []
            for idx in valid_indices:
                if idx in pending_df.index:
                    row = pending_df.loc[idx]
                    if "_复核状态" in row and pd.notna(row["_复核状态"]) and str(row["_复核状态"]).strip() != "":
                        already_reviewed.append(idx)
            
            # 从有效索引中移除已复核的
            valid_indices = [idx for idx in valid_indices if idx not in already_reviewed]
            
            if not valid_indices:
                return False, "所选数据已全部标记为已复核"
            
            reviewed_df = self.load_reviewed_data()
            
            # 提取选中的数据
            selected_data = pending_df.loc[valid_indices].copy()
            selected_data["_复核状态"] = "已复核"
            selected_data["_复核时间"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            selected_data["_复核人员"] = "系统用户"
            
            # 如果提供了修正数据，更新数据
            if correct_data is not None:
                for idx, row in correct_data.iterrows():
                    if idx in selected_data.index:
                        for col in correct_data.columns:
                            if col in selected_data.columns:
                                selected_data.at[idx, col] = row[col]
            
            # 添加到已复核数据
            if reviewed_df.empty:
                reviewed_df = selected_data
            else:
                reviewed_df = pd.concat([reviewed_df, selected_data], ignore_index=True)
            
            # 保存已复核数据
            write_results_to_excel(reviewed_df, self.reviewed_file)
            
            # 从待复核数据中移除
            remaining_df = pending_df.drop(valid_indices).reset_index(drop=True)
            if remaining_df.empty:
                os.remove(self.pending_file)
            else:
                # 保存更新后的待复核数据
                write_results_to_excel(remaining_df, self.pending_file)
            
            # 更新配置
            self.config["total_reviewed"] = self.config.get("total_reviewed", 0) + len(valid_indices)
            self.config["last_review_date"] = datetime.now().strftime("%Y-%m-%d")
            self.save_config()
            
            return True, f"已标记 {len(valid_indices)} 条数据为已复核"
            
        except Exception as e:
            return False, f"标记复核状态失败：{str(e)}"
    
    def append_to_history(self, history_excel_path, force_retrain=False, remove_confidence_column=True, reviewer_name="系统用户"):
        """将已复核数据追加到历史数据"""
        try:
            reviewed_df = self.load_reviewed_data()
            if reviewed_df.empty:
                return False, "无已复核数据可追加"
            
            # 读取历史数据
            if os.path.exists(history_excel_path):
                # 修复：指定引擎
                try:
                    history_df = pd.read_excel(history_excel_path, dtype=str, engine='openpyxl')
                except Exception:
                    try:
                        history_df = pd.read_excel(history_excel_path, dtype=str, engine='xlrd')
                    except Exception as e:
                        return False, f"无法读取历史数据文件：{str(e)}"
            else:
                history_df = pd.DataFrame()
            
            # 移除辅助列
            columns_to_remove = [col for col in reviewed_df.columns if col.startswith('_')]
            
            # 如果需要移除置信度列
            if remove_confidence_column and "置信度" in reviewed_df.columns:
                columns_to_remove.append("置信度")
            
            clean_reviewed_df = reviewed_df.drop(columns=columns_to_remove, errors='ignore')
            
            # 为数据添加追加标记（避免重复追加）
            clean_reviewed_df["_追加时间"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            clean_reviewed_df["_追加人员"] = reviewer_name
            
            # 追加数据
            updated_history_df = pd.concat([history_df, clean_reviewed_df], ignore_index=True)
            
            # 去重（基于所有非系统列）
            non_system_cols = [col for col in clean_reviewed_df.columns if not col.startswith('_')]
            updated_history_df = updated_history_df.drop_duplicates(subset=non_system_cols, keep='first')
            
            # 保存更新后的历史数据
            write_results_to_excel(updated_history_df, history_excel_path)
            
            # 更新配置
            added_count = len(clean_reviewed_df)
            self.config["total_added"] = self.config.get("total_added", 0) + added_count
            self.config["last_append_date"] = datetime.now().strftime("%Y-%m-%d")
            
            # 记录已追加的数据ID（用于避免重复追加）
            if "appended_records" not in self.config:
                self.config["appended_records"] = []
            
            # 获取已追加数据的特征哈希值（用于标识）
            for _, row in clean_reviewed_df.iterrows():
                # 创建数据的唯一标识
                data_id = self._create_data_id(row, non_system_cols)
                if data_id not in self.config["appended_records"]:
                    self.config["appended_records"].append(data_id)
            
            # 清空已复核数据（追加后自动清空）
            if os.path.exists(self.reviewed_file):
                os.remove(self.reviewed_file)
            
            self.save_config()
            
            message = f"✅ 已追加 {added_count} 条数据到历史数据"
            if remove_confidence_column:
                message += "\n⚠️ 已移除置信度列，避免污染历史数据"
            if force_retrain:
                message += "\n⚠️ 建议重新训练模型以包含新数据"
            
            return True, message
            
        except Exception as e:
            return False, f"追加数据失败：{str(e)}"
    
    def _create_data_id(self, row, columns):
        """创建数据的唯一标识符"""
        # 使用特征列的值创建哈希值
        data_str = ""
        for col in columns:
            if col in row:
                data_str += f"{col}:{str(row[col])};"
        return hashlib.md5(data_str.encode('utf-8')).hexdigest()
    
    def clear_reviewed_data(self):
        """清空已复核数据（手动调用）"""
        try:
            if os.path.exists(self.reviewed_file):
                os.remove(self.reviewed_file)
                return True, "已复核数据已清空"
            return False, "无已复核数据可清空"
        except Exception as e:
            return False, f"清空已复核数据失败：{str(e)}"

# ==================== 日志管理器类 ====================
class OperationLogger:
    """日志管理类"""
    
    def __init__(self):
        self.log_json_path = LOG_JSON_PATH
        self.log_txt_path = LOG_TXT_PATH
        self.init_log()
    
    def init_log(self):
        """初始化日志文件"""
        if not os.path.exists(self.log_json_path):
            with open(self.log_json_path, "w", encoding="utf-8") as f:
                json.dump([], f, ensure_ascii=False, indent=4)
        
        if not os.path.exists(self.log_txt_path):
            # 使用utf-8-sig编码写入BOM，确保Windows记事本能正确识别编码
            with open(self.log_txt_path, "w", encoding="utf-8-sig") as f:
                f.write("智能标签填充操作日志\n")
                f.write("="*80 + "\n")
    
    def write_log(self, feat_cols, target_cols, train_data_count, 
                  predict_data_count, confidence_stats=None, mapping_config=None):
        """写入操作日志"""
        try:
            # 读取现有日志
            with open(self.log_json_path, "r", encoding="utf-8") as f:
                logs = json.load(f)
            
            # 构建日志条目
            log_entry = {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "feature_columns": feat_cols,
                "target_columns": target_cols,
                "mapping_config": mapping_config if mapping_config else [],
                "train_data_rows": int(train_data_count),
                "predict_data_rows": int(predict_data_count),
                "confidence_statistics": confidence_stats if confidence_stats else {
                    "mean": 0.0, "min": 0.0, "low_confidence_count": 0
                }
            }
            
            # 写入JSON日志
            logs.append(log_entry)
            with open(self.log_json_path, "w", encoding="utf-8") as f:
                json.dump(logs, f, ensure_ascii=False, indent=4)
            
            # 写入文本日志（使用utf-8-sig编码）
            with open(self.log_txt_path, "a", encoding="utf-8-sig") as f:
                f.write(f"\n【操作时间】：{log_entry['timestamp']}\n")
                f.write(f"【特征列】：{', '.join(log_entry['feature_columns'])}\n")
                f.write(f"【目标列】：{', '.join(log_entry['target_columns'])}\n")
                if mapping_config and len(mapping_config) > 0:
                    f.write(f"【映射规则】：共 {len(mapping_config)} 条\n")
                    for i, (source_file, key_cols, value_col) in enumerate(mapping_config):
                        source_name = os.path.basename(source_file)
                        f.write(f"  规则{i+1}: {source_name} | {key_cols} → {value_col}\n")
                f.write(f"【训练数据行数】：{log_entry['train_data_rows']}\n")
                f.write(f"【预测数据行数】：{log_entry['predict_data_rows']}\n")
                f.write(f"【平均置信度】：{log_entry['confidence_statistics'].get('mean', 0.0):.4f}\n")
                f.write(f"【最低置信度】：{log_entry['confidence_statistics'].get('min', 0.0):.4f}\n")
                f.write(f"【低置信度数据数】：{log_entry['confidence_statistics'].get('low_confidence_count', 0)}\n")
                f.write("-"*80 + "\n")
            
            print(f"✅ 日志写入成功：{log_entry['timestamp']}")
            return True
        except Exception as e:
            print(f"⚠️ 日志写入失败：{str(e)}")
            return False
    
    def show_log_gui(self):
        """显示日志窗口"""
        try:
            with open(self.log_json_path, "r", encoding="utf-8") as f:
                logs = json.load(f)
            
            log_window = tk.Toplevel()
            log_window.title("操作日志查看器")
            log_window.geometry("1000x700")
            
            # 创建Notebook
            notebook = ttk.Notebook(log_window)
            notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # JSON视图
            json_frame = ttk.Frame(notebook)
            notebook.add(json_frame, text="JSON视图")
            
            json_text = scrolledtext.ScrolledText(json_frame, wrap=tk.WORD, font=("Consolas", 10))
            json_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            if logs:
                json_text.insert(tk.END, json.dumps(logs, ensure_ascii=False, indent=2))
            else:
                json_text.insert(tk.END, "暂无操作记录")
            json_text.config(state=tk.DISABLED)
            
            # 表格视图
            table_frame = ttk.Frame(notebook)
            notebook.add(table_frame, text="表格视图")
            
            # 创建Treeview
            columns = ("时间", "特征列数", "目标列数", "训练行数", "预测行数", "平均置信度")
            tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)
            
            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=120)
            
            tree.column("时间", width=180)
            
            # 填充数据
            if logs:
                for log in reversed(logs[-50:]):  # 显示最近50条
                    tree.insert("", tk.END, values=(
                        log['timestamp'],
                        len(log['feature_columns']),
                        len(log['target_columns']),
                        log['train_data_rows'],
                        log['predict_data_rows'],
                        f"{log['confidence_statistics'].get('mean', 0):.4f}"
                    ))
            
            vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
            hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            
            tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")
            
            table_frame.grid_rowconfigure(0, weight=1)
            table_frame.grid_columnconfigure(0, weight=1)
            
            # 统计视图
            stat_frame = ttk.Frame(notebook)
            notebook.add(stat_frame, text="统计信息")
            
            stat_text = scrolledtext.ScrolledText(stat_frame, wrap=tk.WORD, font=("微软雅黑", 10))
            stat_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            if logs:
                total_ops = len(logs)
                recent_10 = logs[-10:] if len(logs) >= 10 else logs
                avg_conf = np.mean([log['confidence_statistics'].get('mean', 0) for log in recent_10])
                total_predict = sum([log['predict_data_rows'] for log in logs])
                
                stat_text.insert(tk.END, f"📊 日志统计信息\n")
                stat_text.insert(tk.END, f"{'='*50}\n")
                stat_text.insert(tk.END, f"总操作次数：{total_ops} 次\n")
                stat_text.insert(tk.END, f"总预测数据量：{total_predict} 行\n")
                stat_text.insert(tk.END, f"最近10次平均置信度：{avg_conf:.4f}\n")
                stat_text.insert(tk.END, f"最近操作时间：{logs[-1]['timestamp']}\n")
                stat_text.insert(tk.END, f"{'='*50}\n\n")
                
                # 按目标列统计
                target_counts = {}
                for log in logs:
                    for target in log['target_columns']:
                        target_counts[target] = target_counts.get(target, 0) + 1
                
                if target_counts:
                    stat_text.insert(tk.END, "🎯 目标列使用频率：\n")
                    for target, count in sorted(target_counts.items(), key=lambda x: x[1], reverse=True):
                        stat_text.insert(tk.END, f"  {target}: {count} 次\n")
            
            stat_text.config(state=tk.DISABLED)
            
            log_window.mainloop()
            
        except Exception as e:
            messagebox.showerror("日志查看错误", f"打开日志失败：{str(e)}")

# ==================== 低置信筛选器类 ====================
class LowConfidenceFilter:
    """低置信数据筛选器"""
    
    def __init__(self, threshold=CONFIDENCE_THRESHOLD):
        self.threshold = threshold
        self.low_conf_data = None
    
    def filter_low_confidence(self, input_path=None, output_path=None):
        """筛选低置信数据"""
        try:
            input_path = input_path or NEW_EXCEL
            output_path = output_path or LOW_CONF_OUTPUT
            
            if not os.path.exists(input_path):
                raise FileNotFoundError(f"未找到文件：{input_path}")
            
            # 修复：指定引擎
            try:
                df = pd.read_excel(input_path, dtype=str, engine='openpyxl')
            except Exception:
                try:
                    df = pd.read_excel(input_path, dtype=str, engine='xlrd')
                except Exception as e:
                    raise ValueError(f"无法读取输入文件：{str(e)}")
            
            if "置信度" not in df.columns:
                raise ValueError("数据文件中无【置信度】列，请先运行主程序进行预测")
            
            # 筛选并排序
            self.low_conf_data = df[pd.to_numeric(df["置信度"], errors='coerce') < self.threshold].copy()
            
            if not self.low_conf_data.empty:
                # 按置信度升序排序
                self.low_conf_data["置信度_numeric"] = pd.to_numeric(self.low_conf_data["置信度"], errors='coerce')
                self.low_conf_data = self.low_conf_data.sort_values(by="置信度_numeric", ascending=True)
                self.low_conf_data = self.low_conf_data.drop(columns=["置信度_numeric"])
                
                # 保存到Excel
                write_results_to_excel(self.low_conf_data, output_path)
                
                # 添加统计信息
                stats = self._get_statistics()
                return True, stats
            else:
                return False, {"count": 0, "min_confidence": None, "output_path": None}
                
        except Exception as e:
            return False, {"error": str(e)}
    
    def _get_statistics(self):
        """获取统计信息"""
        if self.low_conf_data is None or self.low_conf_data.empty:
            return {"count": 0, "min_confidence": None}
        
        return {
            "count": len(self.low_conf_data),
            "min_confidence": float(pd.to_numeric(self.low_conf_data["置信度"], errors='coerce').min()),
            "max_confidence": float(pd.to_numeric(self.low_conf_data["置信度"], errors='coerce').max()),
            "avg_confidence": float(pd.to_numeric(self.low_conf_data["置信度"], errors='coerce').mean()),
            "output_path": LOW_CONF_OUTPUT
        }

# ==================== 版本管理器类 ====================
class VersionManager:
    """模型+映射规则+参数版本管理"""
    
    def __init__(self, model_dir, mapping_dir):
        self.model_dir = model_dir
        self.mapping_dir = mapping_dir
        self.model_backup_dir = os.path.join(model_dir, "backup")
        self.mapping_backup_dir = os.path.join(mapping_dir, "backup")
        self.current_model_path = os.path.join(model_dir, "current_model.pkl")
        self.current_mapping_path = os.path.join(mapping_dir, "current_mapping.pkl")
        
        for folder in [self.model_dir, self.model_backup_dir, self.mapping_dir, self.mapping_backup_dir]:
            if not os.path.exists(folder):
                os.makedirs(folder)
    
    def _get_version_name(self):
        """生成时间戳版本名"""
        return datetime.now().strftime("%Y%m%d_%H%M%S")
    
    def backup_model_and_mapping(self, model_components, mapping_config, model_params):
        """备份模型、映射规则和参数配置"""
        version = self._get_version_name()
        
        # 备份模型 - 包含所有组件：encoder, tfidf_dict, count_dict, classifier
        model_backup_path = os.path.join(self.model_backup_dir, f"model_{version}.pkl")
        save_model_data = {
            "version": version,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "components": model_components,  # (encoder, tfidf_dict, count_dict, classifier)
            "params": model_params,  # 包含随机森林参数和TF-IDF参数
            "mapping_config": mapping_config  # 同时保存映射配置
        }
        joblib.dump(save_model_data, model_backup_path)
        joblib.dump(save_model_data, self.current_model_path)
        
        # 备份映射规则（独立备份）
        mapping_backup_path = os.path.join(self.mapping_backup_dir, f"mapping_{version}.pkl")
        save_mapping_data = {
            "version": version,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "mapping_config": mapping_config
        }
        joblib.dump(save_mapping_data, mapping_backup_path)
        joblib.dump(save_mapping_data, self.current_mapping_path)
        
        print(f"💾 模型+映射规则+参数配置备份完成（版本：{version}）")
        return version
    
    def get_available_models(self):
        """获取所有可用的模型备份"""
        models = []
        
        # 检查当前模型
        if os.path.exists(self.current_model_path):
            try:
                model_data = joblib.load(self.current_model_path)
                if "components" in model_data:
                    models.append({
                        "path": self.current_model_path,
                        "version": model_data.get("version", "当前模型"),
                        "timestamp": model_data.get("timestamp", "未知时间"),
                        "params": model_data.get("params", {}),
                        "mapping_config": model_data.get("mapping_config", [])
                    })
            except Exception as e:
                print(f"加载当前模型失败：{str(e)}")
        
        # 检查备份模型
        if os.path.exists(self.model_backup_dir):
            for filename in os.listdir(self.model_backup_dir):
                if filename.endswith(".pkl") and filename.startswith("model_"):
                    model_path = os.path.join(self.model_backup_dir, filename)
                    try:
                        model_data = joblib.load(model_path)
                        if "components" in model_data:
                            models.append({
                                "path": model_path,
                                "version": model_data.get("version", filename.replace("model_", "").replace(".pkl", "")),
                                "timestamp": model_data.get("timestamp", "未知时间"),
                                "params": model_data.get("params", {}),
                                "mapping_config": model_data.get("mapping_config", [])
                            })
                    except Exception as e:
                        print(f"加载备份模型 {filename} 失败：{str(e)}")
        
        # 按时间戳排序，最新的在前面
        models.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
        return models
    
    def load_model(self, model_path):
        """加载指定路径的模型"""
        try:
            if not os.path.exists(model_path):
                raise FileNotFoundError(f"模型文件不存在：{model_path}")
            
            model_data = joblib.load(model_path)
            
            if "components" not in model_data:
                raise ValueError("模型文件格式不正确，缺少components字段")
            
            # 提取组件
            encoder = model_data["components"][0]
            tfidf_dict = model_data["components"][1]
            count_dict = model_data["components"][2]
            classifier = model_data["components"][3]
            
            # 提取参数
            params = model_data.get("params", {})
            
            # 提取映射配置
            mapping_config = model_data.get("mapping_config", [])
            
            print(f"✅ 模型加载成功：{model_data.get('version', '未知版本')}")
            return encoder, tfidf_dict, count_dict, classifier, params, mapping_config
            
        except Exception as e:
            raise Exception(f"加载模型失败：{str(e)}")

# ==================== 进度窗口类 ====================
class ProgressWindow:
    """进度显示窗口"""
    
    def __init__(self, parent, title="处理中"):
        self.window = tk.Toplevel(parent)
        self.window.title(title)
        self.window.geometry("400x150")
        self.window.resizable(False, False)
        
        # 居中显示
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - 200
        y = (self.window.winfo_screenheight() // 2) - 75
        self.window.geometry(f"400x150+{x}+{y}")
        
        # 设置窗口关闭协议，防止意外关闭
        self.window.protocol("WM_DELETE_WINDOW", self.on_close)
        
        # 状态标签
        self.status_label = ttk.Label(self.window, text="初始化...", font=("宋体", 10))
        self.status_label.pack(pady=10)
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            self.window,
            variable=self.progress_var,
            maximum=100,
            length=350
        )
        self.progress_bar.pack(pady=10)
        
        # 百分比标签
        self.percent_label = ttk.Label(self.window, text="0%", font=("宋体", 9))
        self.percent_label.pack(pady=5)
        
        # 标记窗口是否已销毁
        self.is_closed = False
        
        self.window.update()
    
    def update(self, progress, status):
        """更新进度"""
        if not self.is_closed:
            try:
                self.progress_var.set(progress)
                self.percent_label.config(text=f"{progress:.0f}%")
                self.status_label.config(text=status)
                self.window.update_idletasks()
            except tk.TclError:
                # 窗口可能已经被销毁，忽略错误
                self.is_closed = True
    
    def close(self):
        """安全关闭进度窗口"""
        if not self.is_closed:
            try:
                self.is_closed = True
                self.window.destroy()
            except:
                pass
    
    def on_close(self):
        """窗口关闭事件处理"""
        # 不允许用户手动关闭进度窗口
        pass

# ==================== 历史模型选择窗口类 ====================
class ModelSelectionWindow:
    """历史模型选择窗口"""
    
    def __init__(self, parent, version_manager, title="选择历史模型"):
        self.window = tk.Toplevel(parent)
        self.window.title(title)
        self.window.geometry("800x500")
        self.window.resizable(True, True)
        
        # 居中显示
        self.window.update_idletasks()
        x = (parent.winfo_screenwidth() // 2) - 400
        y = (parent.winfo_screenheight() // 2) - 250
        self.window.geometry(f"800x500+{x}+{y}")
        
        self.version_manager = version_manager
        self.selected_model = None
        
        # 创建界面
        self.create_ui()
        
    def create_ui(self):
        """创建选择界面"""
        # 标题
        title_frame = ttk.Frame(self.window)
        title_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(title_frame, text="请选择要使用的历史模型", font=("微软雅黑", 12, "bold")).pack()
        
        # 模型列表
        list_frame = ttk.LabelFrame(self.window, text="可用模型列表")
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 创建Treeview
        columns = ("版本", "时间", "特征列数", "目标列数", "映射规则数", "随机森林树数")
        self.tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=15)
        
        for col in columns:
            self.tree.heading(col, text=col)
        
        self.tree.column("版本", width=100)
        self.tree.column("时间", width=150)
        self.tree.column("特征列数", width=80)
        self.tree.column("目标列数", width=80)
        self.tree.column("映射规则数", width=80)
        self.tree.column("随机森林树数", width=100)
        
        # 添加滚动条
        vsb = ttk.Scrollbar(list_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(list_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        list_frame.grid_rowconfigure(0, weight=1)
        list_frame.grid_columnconfigure(0, weight=1)
        
        # 详细信息区域
        detail_frame = ttk.LabelFrame(self.window, text="模型详细信息")
        detail_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.detail_text = scrolledtext.ScrolledText(detail_frame, height=8, wrap=tk.WORD, font=("Consolas", 9))
        self.detail_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.detail_text.config(state=tk.DISABLED)
        
        # 按钮区域
        button_frame = ttk.Frame(self.window)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(button_frame, text="刷新列表", command=self.refresh_list).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="选择模型", command=self.select_model, 
                  style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="取消", command=self.window.destroy).pack(side=tk.RIGHT, padx=5)
        
        # 绑定选择事件
        self.tree.bind("<<TreeviewSelect>>", self.on_model_selected)
        
        # 初始加载模型列表
        self.refresh_list()
    
    def refresh_list(self):
        """刷新模型列表"""
        # 清空现有数据
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # 获取可用模型
        models = self.version_manager.get_available_models()
        
        if not models:
            self.detail_text.config(state=tk.NORMAL)
            self.detail_text.delete(1.0, tk.END)
            self.detail_text.insert(tk.END, "未找到可用的历史模型")
            self.detail_text.config(state=tk.DISABLED)
            return
        
        # 添加模型到列表
        for model in models:
            params = model.get("params", {})
            rf_params = params.get("rf", {})
            mapping_config = model.get("mapping_config", [])
            
            # 从模型数据中提取特征列和目标列信息
            feat_cols = []
            target_cols = []
            if "components" in model:
                # 这里可以根据需要从模型中提取更多信息
                pass
            
            self.tree.insert("", tk.END, values=(
                model.get("version", "未知"),
                model.get("timestamp", "未知时间"),
                len(feat_cols) if feat_cols else "N/A",
                len(target_cols) if target_cols else "N/A",
                len(mapping_config),
                rf_params.get("n_estimators", "N/A")
            ), tags=("model",))
        
        # 默认选择第一个
        if self.tree.get_children():
            first_item = self.tree.get_children()[0]
            self.tree.selection_set(first_item)
            self.tree.focus(first_item)
            self.update_detail_view(models[0])
    
    def on_model_selected(self, event):
        """模型选择事件处理"""
        selection = self.tree.selection()
        if not selection:
            return
        
        # 获取选中的模型索引
        item = selection[0]
        index = self.tree.index(item)
        
        # 获取模型列表
        models = self.version_manager.get_available_models()
        if index < len(models):
            self.update_detail_view(models[index])
    
    def update_detail_view(self, model):
        """更新详细视图"""
        params = model.get("params", {})
        rf_params = params.get("rf", {})
        tfidf_params = params.get("tfidf", {})
        mapping_config = model.get("mapping_config", [])
        
        detail_text = f"模型版本: {model.get('version', '未知')}\n"
        detail_text += f"创建时间: {model.get('timestamp', '未知时间')}\n"
        detail_text += f"模型路径: {model.get('path', '未知')}\n\n"
        
        detail_text += "随机森林参数:\n"
        detail_text += f"  决策树数量: {rf_params.get('n_estimators', 'N/A')}\n"
        detail_text += f"  最大深度: {rf_params.get('max_depth', '不限制')}\n"
        detail_text += f"  最小分裂样本数: {rf_params.get('min_samples_split', 2)}\n"
        detail_text += f"  最小叶子样本数: {rf_params.get('min_samples_leaf', 1)}\n\n"
        
        detail_text += "TF-IDF参数:\n"
        detail_text += f"  n-gram范围: {tfidf_params.get('ngram_range', (1, 2))}\n"
        detail_text += f"  最大特征数: {tfidf_params.get('max_features', 100)}\n"
        detail_text += f"  停用词: {tfidf_params.get('stop_words', 'english')}\n\n"
        
        detail_text += f"映射规则数量: {len(mapping_config)}\n"
        if mapping_config:
            detail_text += "映射规则详情:\n"
            for i, (source_file, key_cols, value_col) in enumerate(mapping_config):
                source_name = os.path.basename(source_file)
                detail_text += f"  规则{i+1}: {source_name} | {key_cols} → {value_col}\n"
        
        self.detail_text.config(state=tk.NORMAL)
        self.detail_text.delete(1.0, tk.END)
        self.detail_text.insert(tk.END, detail_text)
        self.detail_text.config(state=tk.DISABLED)
    
    def select_model(self):
        """选择模型"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请先选择一个模型")
            return
        
        # 获取选中的模型索引
        item = selection[0]
        index = self.tree.index(item)
        
        # 获取模型列表
        models = self.version_manager.get_available_models()
        if index < len(models):
            self.selected_model = models[index]
            self.window.destroy()
        else:
            messagebox.showerror("错误", "选择的模型不存在")
    
    def get_selected_model(self):
        """获取选中的模型"""
        return self.selected_model

# ==================== 增强版：人工复核管理界面（修复版）====================
class HumanReviewWindow:
    """人工复核管理窗口 - 修复版：解决全选和重复数据问题"""
    
    def __init__(self, parent, incremental_learner):
        self.window = tk.Toplevel(parent)
        self.window.title("人工复核管理")
        self.window.geometry("1000x700")
        
        self.incremental_learner = incremental_learner
        self.pending_data_cache = None  # 缓存待复核数据
        self.selected_items = set()  # 存储选中的项目
        self.all_items_selected = False  # 标记是否全选状态
        
        # 居中显示
        self.window.update_idletasks()
        x = (parent.winfo_screenwidth() // 2) - 500
        y = (parent.winfo_screenheight() // 2) - 350
        self.window.geometry(f"1000x700+{x}+{y}")
        
        self.create_ui()
        self.load_data()
    
    def create_ui(self):
        """创建界面"""
        # 创建Notebook
        notebook = ttk.Notebook(self.window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 1. 待复核数据页
        pending_frame = ttk.Frame(notebook)
        notebook.add(pending_frame, text="待复核数据")
        self.create_pending_tab(pending_frame)
        
        # 2. 已复核数据页
        reviewed_frame = ttk.Frame(notebook)
        notebook.add(reviewed_frame, text="已复核数据")
        self.create_reviewed_tab(reviewed_frame)
        
        # 3. 统计信息页
        stats_frame = ttk.Frame(notebook)
        notebook.add(stats_frame, text="统计信息")
        self.create_stats_tab(stats_frame)
        
        # 4. 配置页（新增）
        config_frame = ttk.Frame(notebook)
        notebook.add(config_frame, text="配置")
        self.create_config_tab(config_frame)
    
    def create_pending_tab(self, parent):
        """创建待复核数据页 - 修复全选功能"""
        # 工具栏
        toolbar = ttk.Frame(parent)
        toolbar.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(toolbar, text="刷新数据", command=self.load_data).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="全选/取消全选", command=self.toggle_select_all).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="标记为正确", command=self.mark_as_correct).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="编辑数据", command=self.edit_data).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="导出选中", command=self.export_selected).pack(side=tk.LEFT, padx=2)
        
        # 数据表格
        table_frame = ttk.Frame(parent)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 创建Treeview
        columns = ["选择", "复核原因", "置信度", "数据来源", "时间戳"]  # 动态列会在load_data中添加
        self.pending_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20, selectmode='extended')
        
        # 添加复选框列
        self.pending_tree.heading("选择", text="选择")
        self.pending_tree.column("选择", width=50)
        
        # 添加其他固定列
        for col in columns[1:]:
            self.pending_tree.heading(col, text=col)
            self.pending_tree.column(col, width=100)
        
        # 滚动条
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.pending_tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.pending_tree.xview)
        self.pending_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.pending_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        # 绑定双击事件查看详情
        self.pending_tree.bind("<Double-1>", self.show_detail)
        
        # 状态信息
        self.pending_status = ttk.Label(parent, text="共 0 条待复核数据")
        self.pending_status.pack(pady=5)
    
    def create_reviewed_tab(self, parent):
        """创建已复核数据页"""
        # 工具栏
        toolbar = ttk.Frame(parent)
        toolbar.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(toolbar, text="刷新数据", command=self.load_data).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="追加到历史数据", command=self.append_to_history).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="导出全部", command=self.export_reviewed).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="清空已复核", command=self.clear_reviewed).pack(side=tk.LEFT, padx=2)
        
        # 数据表格
        table_frame = ttk.Frame(parent)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        columns = ["复核状态", "复核时间", "复核原因", "数据来源", "复核人员"]
        self.reviewed_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)
        
        for col in columns:
            self.reviewed_tree.heading(col, text=col)
            self.reviewed_tree.column(col, width=120)
        
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.reviewed_tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.reviewed_tree.xview)
        self.reviewed_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.reviewed_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        # 状态信息
        self.reviewed_status = ttk.Label(parent, text="共 0 条已复核数据")
        self.reviewed_status.pack(pady=5)
    
    def create_stats_tab(self, parent):
        """创建统计信息页"""
        # 统计信息显示
        stats_text = scrolledtext.ScrolledText(parent, wrap=tk.WORD, font=("微软雅黑", 10))
        stats_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.stats_text = stats_text
    
    def create_config_tab(self, parent):
        """创建配置页"""
        config_frame = ttk.Frame(parent)
        config_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 保护已编辑数据选项
        protect_frame = ttk.LabelFrame(config_frame, text="数据保护设置")
        protect_frame.pack(fill=tk.X, pady=10)
        
        self.protect_edited_var = tk.BooleanVar(
            value=self.incremental_learner.config.get("preserve_edited_data", True))
        
        ttk.Checkbutton(protect_frame, text="保护已编辑数据（手动提取时不会覆盖已标记的数据）", 
                       variable=self.protect_edited_var,
                       command=self.save_protect_setting).pack(padx=10, pady=10)
        
        # 说明文本
        info_frame = ttk.LabelFrame(config_frame, text="说明")
        info_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        info_text = scrolledtext.ScrolledText(info_frame, wrap=tk.WORD, font=("微软雅黑", 9), height=10)
        info_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        info = """使用说明：

1. 全选功能：点击"全选/取消全选"按钮可以快速选择或取消所有数据
2. 数据保护：启用"保护已编辑数据"后，手动提取增量数据时不会覆盖已经标记为已复核或已编辑的数据
3. 编辑数据：可以直接在Excel中编辑"增量数据_待复核.xlsx"文件，编辑后保存即可
4. 标记数据：在界面上选择数据后点击"标记为正确"，数据会移动到已复核列表
5. 追加数据：已复核数据可以追加到历史数据，但会移除置信度列以避免污染历史数据
6. 清空已复核：追加后可以清空已复核数据，避免重复追加

注意事项：
- 历史数据追加时会自动移除重复数据
- 身份证、手机号等长数字会保持文本格式，避免科学计数法
- 建议在追加数据后重新训练模型以获得最佳效果
"""
        
        info_text.insert(tk.END, info)
        info_text.config(state=tk.DISABLED)
    
    def load_data(self):
        """加载数据 - 修复：过滤已标记为已复核的数据"""
        try:
            # 加载待复核数据
            pending_df = self.incremental_learner.load_pending_data()
            
            # 过滤掉已经标记为已复核的数据（只显示未复核的）
            if not pending_df.empty and "_复核状态" in pending_df.columns:
                # 只显示未复核的数据
                pending_df = pending_df[pending_df["_复核状态"].isna() | (pending_df["_复核状态"] == "")]
            
            self.pending_data_cache = pending_df  # 缓存数据
            
            # 清空现有数据
            for item in self.pending_tree.get_children():
                self.pending_tree.delete(item)
            
            if not pending_df.empty:
                # 动态添加数据列（排除辅助列）
                data_cols = [col for col in pending_df.columns if not col.startswith('_')]
                
                # 更新列配置
                current_columns = list(self.pending_tree["columns"])
                new_columns = ["选择"] + data_cols + ["_复核原因", "_数据来源", "_时间戳"]
                
                if set(current_columns) != set(new_columns):
                    # 重新配置Treeview
                    self.pending_tree.destroy()
                    
                    # 创建新的Treeview
                    table_frame = self.pending_tree.master
                    self.pending_tree = ttk.Treeview(table_frame, columns=new_columns, show="headings", height=20, selectmode='extended')
                    
                    # 配置列
                    for col in new_columns:
                        if col == "选择":
                            width = 50
                        elif col in ["_复核原因", "_数据来源", "_时间戳"]:
                            width = 150
                        else:
                            width = 100
                        
                        display_name = col.replace('_', '') if col.startswith('_') else col
                        self.pending_tree.heading(col, text=display_name)
                        self.pending_tree.column(col, width=width)
                    
                    # 重新配置滚动条
                    vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.pending_tree.yview)
                    hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.pending_tree.xview)
                    self.pending_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
                    
                    self.pending_tree.grid(row=0, column=0, sticky="nsew")
                    vsb.grid(row=0, column=1, sticky="ns")
                    hsb.grid(row=1, column=0, sticky="ew")
                    
                    # 重新绑定事件
                    self.pending_tree.bind("<Double-1>", self.show_detail)
                
                # 清空选中集合
                self.selected_items.clear()
                self.all_items_selected = False
                
                # 添加数据
                for idx, row in pending_df.iterrows():
                    # 所有数据都显示为未选择状态
                    checkbox_state = "□"
                    
                    values = [checkbox_state] + [str(row.get(col, "")) for col in data_cols] + \
                            [str(row.get("_复核原因", "")), 
                             str(row.get("_数据来源", "")), 
                             str(row.get("_时间戳", ""))]
                    
                    item_id = str(idx)
                    self.pending_tree.insert("", tk.END, values=values, iid=item_id)
                
                self.pending_status.config(text=f"共 {len(pending_df)} 条待复核数据")
            else:
                self.pending_status.config(text="暂无待复核数据")
            
            # 加载已复核数据
            reviewed_df = self.incremental_learner.load_reviewed_data()
            
            # 清空现有数据
            for item in self.reviewed_tree.get_children():
                self.reviewed_tree.delete(item)
            
            if not reviewed_df.empty:
                for idx, row in reviewed_df.iterrows():
                    values = [
                        str(row.get("_复核状态", "")),
                        str(row.get("_复核时间", "")),
                        str(row.get("_复核原因", "")),
                        str(row.get("_数据来源", "")),
                        str(row.get("_复核人员", ""))
                    ]
                    self.reviewed_tree.insert("", tk.END, values=values)
                
                self.reviewed_status.config(text=f"共 {len(reviewed_df)} 条已复核数据")
            else:
                self.reviewed_status.config(text="暂无已复核数据")
            
            # 更新统计信息
            self.update_stats()
            
        except Exception as e:
            messagebox.showerror("加载失败", f"加载数据失败：{str(e)}")
    
    def update_stats(self):
        """更新统计信息"""
        try:
            config = self.incremental_learner.config
            
            stats_text = f"📊 增量学习统计信息\n"
            stats_text += "="*50 + "\n"
            stats_text += f"累计复核数据：{config.get('total_reviewed', 0)} 条\n"
            stats_text += f"累计追加数据：{config.get('total_added', 0)} 条\n"
            stats_text += f"最后复核日期：{config.get('last_review_date', '从未复核')}\n"
            stats_text += f"最后追加日期：{config.get('last_append_date', '从未追加')}\n"
            stats_text += f"自动增量学习：{'开启' if config.get('auto_incremental_learning', True) else '关闭'}\n"
            stats_text += f"置信度阈值：{config.get('confidence_threshold', CONFIDENCE_THRESHOLD)}\n"
            stats_text += f"最小增量样本：{config.get('min_incremental_samples', 10)} 条\n"
            stats_text += f"保护已编辑数据：{'是' if config.get('preserve_edited_data', True) else '否'}\n"
            stats_text += "="*50 + "\n\n"
            
            # 待复核数据统计
            pending_df = self.incremental_learner.load_pending_data()
            if not pending_df.empty:
                stats_text += "📋 待复核数据统计\n"
                stats_text += f"总记录数：{len(pending_df)} 条\n"
                
                if "_复核原因" in pending_df.columns:
                    reason_counts = pending_df["_复核原因"].value_counts()
                    for reason, count in reason_counts.items():
                        stats_text += f"  {reason}: {count} 条\n"
                
                if "_复核状态" in pending_df.columns:
                    reviewed_count = pending_df["_复核状态"].notna().sum()
                    stats_text += f"  已标记数：{reviewed_count} 条\n"
                    stats_text += f"  待处理数：{len(pending_df) - reviewed_count} 条\n"
                
                stats_text += "\n"
            
            # 已复核数据统计
            reviewed_df = self.incremental_learner.load_reviewed_data()
            if not reviewed_df.empty:
                stats_text += "✅ 已复核数据统计\n"
                stats_text += f"总记录数：{len(reviewed_df)} 条\n"
                
                if "_复核时间" in reviewed_df.columns:
                    reviewed_df["_复核日期"] = reviewed_df["_复核时间"].str[:10]
                    date_counts = reviewed_df["_复核日期"].value_counts().sort_index(ascending=False)
                    for date, count in date_counts.head(5).items():  # 显示最近5天
                        stats_text += f"  • {date}: {count} 条\n"
            
            # 历史数据统计
            try:
                if os.path.exists(HISTORY_EXCEL):
                    history_df = pd.read_excel(HISTORY_EXCEL, dtype=str)
                    stats_text += f"\n📚 历史数据总计：{len(history_df):,} 条\n"
                    
                    # 按目标列统计
                    for target_col in ["标签", "分类", "类别"]:  # 常见的目标列名
                        if target_col in history_df.columns:
                            cat_count = history_df[target_col].nunique()
                            stats_text += f"  • {target_col}: {cat_count} 个类别\n"
            except:
                pass
            
            self.stats_text.config(state=tk.NORMAL)
            self.stats_text.delete(1.0, tk.END)
            self.stats_text.insert(tk.END, stats_text)
            self.stats_text.config(state=tk.DISABLED)
            
        except Exception as e:
            print(f"更新统计信息失败：{str(e)}")
    
    def toggle_select_all(self):
        """全选/取消全选 - 修复版本：确保正确切换状态"""
        try:
            all_items = self.pending_tree.get_children()
            
            if not all_items:
                return
            
            # 切换全选状态
            self.all_items_selected = not self.all_items_selected
            
            if self.all_items_selected:
                # 全选所有项目
                self.pending_tree.selection_set(all_items)
                for item in all_items:
                    current_values = list(self.pending_tree.item(item, "values"))
                    if current_values:
                        current_values[0] = "☑"
                        self.pending_tree.item(item, values=current_values)
                    self.selected_items.add(item)
            else:
                # 取消全选
                self.pending_tree.selection_remove(all_items)
                for item in all_items:
                    current_values = list(self.pending_tree.item(item, "values"))
                    if current_values:
                        current_values[0] = "□"
                        self.pending_tree.item(item, values=current_values)
                self.selected_items.clear()
                
        except Exception as e:
            print(f"全选功能错误：{str(e)}")
    
    def mark_as_correct(self):
        """标记选中数据为正确 - 修复版本：确保正确处理选择状态"""
        try:
            # 获取选中的项目
            selected_items = self.pending_tree.selection()
            if not selected_items:
                messagebox.showwarning("提示", "请先选择要标记的数据")
                return
            
            # 获取有效的索引（确保数据存在）
            indices = []
            for item in selected_items:
                try:
                    # 检查复选框状态
                    values = self.pending_tree.item(item, "values")
                    if values and values[0] == "☑":
                        indices.append(int(item))
                except:
                    continue
            
            if not indices:
                messagebox.showinfo("提示", "请先选择要标记的数据（确保复选框为☑状态）")
                return
            
            success, message = self.incremental_learner.mark_as_reviewed(indices)
            
            if success:
                messagebox.showinfo("成功", message)
                self.load_data()  # 重新加载数据以更新显示
            else:
                messagebox.showerror("失败", message)
                
        except Exception as e:
            messagebox.showerror("标记失败", f"标记数据失败：{str(e)}")
    
    def edit_data(self):
        """编辑选中数据"""
        selected = self.pending_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要编辑的数据")
            return
        
        try:
            # 打开待复核文件
            pending_file = self.incremental_learner.pending_file
            if os.path.exists(pending_file):
                if os.name == 'nt':
                    os.startfile(pending_file)
                else:
                    import subprocess
                    subprocess.run(['open' if sys.platform == 'darwin' else 'xdg-open', pending_file])
                
                messagebox.showinfo("提示", "请在Excel中编辑数据，保存后点击刷新按钮更新显示")
            else:
                messagebox.showwarning("提示", "待复核文件不存在")
                
        except Exception as e:
            messagebox.showerror("打开失败", f"无法打开文件：{str(e)}")
    
    def export_selected(self):
        """导出选中数据"""
        selected = self.pending_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要导出的数据")
            return
        
        try:
            # 选择保存路径
            filename = filedialog.asksaveasfilename(
                title="保存选中数据",
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
            )
            
            if not filename:
                return
            
            # 使用缓存的数据
            if self.pending_data_cache is not None:
                # 提取选中的数据
                selected_indices = [int(item) for item in selected]
                selected_df = self.pending_data_cache.iloc[selected_indices]
                
                # 保存到文件
                write_results_to_excel(selected_df, filename)
                
                messagebox.showinfo("成功", f"已导出 {len(selected_indices)} 条数据到：\n{filename}")
            else:
                messagebox.showerror("错误", "数据缓存为空，请刷新后重试")
            
        except Exception as e:
            messagebox.showerror("导出失败", f"导出数据失败：{str(e)}")
    
    def show_detail(self, event):
        """显示数据详情"""
        item = self.pending_tree.selection()
        if not item:
            return
        
        item = item[0]
        values = self.pending_tree.item(item, "values")
        
        # 创建详情窗口
        detail_window = tk.Toplevel(self.window)
        detail_window.title("数据详情")
        detail_window.geometry("600x400")
        
        # 居中显示
        detail_window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - 300
        y = (self.window.winfo_screenheight() // 2) - 200
        detail_window.geometry(f"600x400+{x}+{y}")
        
        # 创建文本区域
        detail_text = scrolledtext.ScrolledText(detail_window, wrap=tk.WORD, font=("Consolas", 10))
        detail_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 获取列名
        columns = self.pending_tree["columns"]
        
        # 构建详情文本
        detail_content = "📋 数据详情\n"
        detail_content += "="*50 + "\n"
        
        for i, col in enumerate(columns):
            if values[i]:
                detail_content += f"{col}: {values[i]}\n"
        
        detail_text.insert(tk.END, detail_content)
        detail_text.config(state=tk.DISABLED)
    
    def append_to_history(self):
        """将已复核数据追加到历史数据"""
        if messagebox.askyesno("确认", 
            "确定要将已复核数据追加到历史数据吗？\n"
            "注意：追加时会自动移除置信度列，避免污染历史数据。\n"
            "追加后建议重新训练模型。"):
            
            # 询问是否清空已复核数据
            clear_after_append = messagebox.askyesno("清空已复核数据", 
                "追加完成后是否清空已复核数据？\n"
                "建议清空以避免重复追加。")
            
            success, message = self.incremental_learner.append_to_history(
                HISTORY_EXCEL, 
                remove_confidence_column=True,  # 移除置信度列
                reviewer_name="用户复核"
            )
            
            if success:
                messagebox.showinfo("成功", message)
                
                # 如果用户选择清空已复核数据
                if clear_after_append:
                    clear_success, clear_message = self.incremental_learner.clear_reviewed_data()
                    if clear_success:
                        messagebox.showinfo("成功", f"{message}\n{clear_message}")
                
                self.load_data()
                
                # 询问是否重新训练模型
                if messagebox.askyesno("重新训练", "是否要重新训练模型以包含新增数据？"):
                    # 这里需要调用主窗口的重训练方法
                    self.window.master.retrain_with_incremental()
            else:
                messagebox.showerror("失败", message)
    
    def export_reviewed(self):
        """导出已复核数据"""
        try:
            # 选择保存路径
            filename = filedialog.asksaveasfilename(
                title="保存已复核数据",
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
            )
            
            if not filename:
                return
            
            # 加载已复核数据
            reviewed_df = self.incremental_learner.load_reviewed_data()
            
            if reviewed_df.empty:
                messagebox.showwarning("提示", "无已复核数据可导出")
                return
            
            # 保存到文件
            write_results_to_excel(reviewed_df, filename)
            
            messagebox.showinfo("成功", f"已导出 {len(reviewed_df)} 条已复核数据到：\n{filename}")
            
        except Exception as e:
            messagebox.showerror("导出失败", f"导出数据失败：{str(e)}")
    
    def clear_reviewed(self):
        """清空已复核数据"""
        if messagebox.askyesno("确认", "确定要清空所有已复核数据吗？此操作不可恢复！"):
            success, message = self.incremental_learner.clear_reviewed_data()
            if success:
                messagebox.showinfo("成功", message)
                self.load_data()
            else:
                messagebox.showerror("失败", message)
    
    def save_protect_setting(self):
        """保存保护设置"""
        self.incremental_learner.config["preserve_edited_data"] = self.protect_edited_var.get()
        self.incremental_learner.save_config()
        messagebox.showinfo("成功", "保护设置已保存")

# ==================== 主程序类 ====================
class SmartLabelToolkit:
    """智能标签打标工具主程序"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("智能数据打标工具 v3.1 - 修复版")
        self.root.geometry("1200x800")
        
        # 初始化管理器
        self.logger = OperationLogger()
        self.filter = LowConfidenceFilter()
        self.version_manager = VersionManager(MODEL_DIR, MAPPING_DIR)
        # 新增：初始化增量学习管理器
        self.incremental_learner = IncrementalLearner()
        self.data_comparator = None
        
        # 程序状态
        self.is_exiting = False
        
        # 配置变量
        self.feat_cols = []
        self.target_cols = []
        self.text_cols = []
        self.mapping_config = []  # 现在格式为 (source_file, key_cols, value_col)
        self.model_params = {}
        self.current_model = None  # 当前加载的模型
        
        # 加载历史配置
        self.feat_cols, self.target_cols, self.text_cols = load_config()
        
        # 初始化默认参数
        self.init_default_params()
        
        # 修复：确保所有UI元素在__init__中创建
        self._init_ui_elements()
        
        # 设置窗口关闭协议
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # 创建界面
        self.create_ui()
    
    def _init_ui_elements(self):
        """初始化所有UI元素变量，防止EXE打包后属性错误"""
        # 主处理页元素
        self.history_file_status = None
        self.new_file_status = None
        self.status_text = None
        self.progress_var = None
        self.progress_bar = None
        self.start_button = None
        self.model_list_frame = None
        self.model_listbox = None
        
        # 列选择页元素
        self.feat_vars = {}
        self.text_vars = {}
        self.target_vars = {}
        
        # 映射配置页元素
        self.mapping_file_var = None
        self.columns_listbox = None
        self.key_listbox = None
        self.value_listbox = None
        self.rules_listbox = None
        self.current_file_label = None
        
        # 参数配置页元素
        self.n_estimators_var = None
        self.max_depth_var = None
        self.min_samples_split_var = None
        self.min_samples_leaf_var = None
        self.random_state_var = None
        self.ngram_min_var = None
        self.ngram_max_var = None
        self.max_features_var = None
        self.stop_words_var = None
        self.confidence_threshold_var = None
        
        # 模型选择
        self.model_choice_var = None
        
        # 筛选页元素
        self.filter_threshold_var = None
        self.filter_threshold_label = None
        self.filter_input_var = None
        self.filter_result_text = None
        
        # 日志页元素
        self.log_text = None
        
        # 增量学习页元素
        self.incremental_frame = None
        self.auto_incremental_var = None
        self.incremental_threshold_var = None
        self.threshold_label = None
        self.incremental_stats_text = None
        self.protect_edited_var = None
        
        # 状态栏
        self.status_bar = None
        
    def init_default_params(self):
        """初始化默认参数"""
        self.model_params = {
            "rf": {
                "n_estimators": 100,
                "max_depth": None,
                "min_samples_split": 2,
                "min_samples_leaf": 1,
                "random_state": 42,
                "n_jobs": -1
            },
            "tfidf": {
                "ngram_range": (1, 2),
                "max_features": 100,
                "stop_words": "english"
            }
        }
    
    def create_ui(self):
        """创建Notebook界面"""
        # 创建Notebook
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 1. 主处理页
        self.main_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.main_frame, text="🚀 主处理")
        self.create_main_tab()
        
        # 2. 列选择页
        self.columns_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.columns_frame, text="📊 列选择")
        self.create_columns_tab()
        
        # 3. 映射配置页
        self.mapping_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.mapping_frame, text="🗺️ 映射配置")
        self.create_mapping_tab()
        
        # 4. 参数配置页
        self.params_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.params_frame, text="⚙️ 参数配置")
        self.create_params_tab()
        
        # 5. 日志管理页
        self.log_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.log_frame, text="📋 日志管理")
        self.create_log_tab()
        
        # 6. 低置信筛选页
        self.filter_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.filter_frame, text="🔍 低置信筛选")
        self.create_filter_tab()
        
        # 7. 增量学习页（新增）
        self.incremental_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.incremental_frame, text="🔄 增量学习")
        self.create_incremental_tab()
        
        # 状态栏
        self.status_bar = ttk.Label(self.root, text="就绪", relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
    
    def create_main_tab(self):
        """创建主处理页"""
        # 顶部按钮区
        top_frame = ttk.Frame(self.main_frame)
        top_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(top_frame, text="跳转到列选择", 
                  command=lambda: self.notebook.select(1)).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text="跳转到映射配置", 
                  command=lambda: self.notebook.select(2)).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text="跳转到参数配置", 
                  command=lambda: self.notebook.select(3)).pack(side=tk.LEFT, padx=5)
        
        # 文件检查
        file_frame = ttk.LabelFrame(self.main_frame, text="文件检查")
        file_frame.pack(fill=tk.X, padx=10, pady=10)
        
        check_frame = ttk.Frame(file_frame)
        check_frame.pack(fill=tk.X, padx=10, pady=10)
        
        self.history_file_status = ttk.Label(check_frame, text="❌ 历史数据文件未检查")
        self.history_file_status.pack(side=tk.LEFT, padx=5)
        
        self.new_file_status = ttk.Label(check_frame, text="❌ 新数据文件未检查")
        self.new_file_status.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(check_frame, text="检查文件", command=self.check_files).pack(side=tk.RIGHT, padx=5)
        
        # 模型选择
        model_frame = ttk.LabelFrame(self.main_frame, text="模型选择")
        model_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # 单选按钮框架
        radio_frame = ttk.Frame(model_frame)
        radio_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.model_choice_var = tk.StringVar(value="train_new")
        ttk.Radiobutton(radio_frame, text="训练新模型", variable=self.model_choice_var, 
                       value="train_new", command=self.on_model_choice_changed).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(radio_frame, text="使用历史模型", variable=self.model_choice_var, 
                       value="use_existing", command=self.on_model_choice_changed).pack(side=tk.LEFT, padx=10)
        
        # 历史模型选择列表（默认隐藏）
        self.model_list_frame = ttk.Frame(model_frame)
        self.model_list_frame.pack(fill=tk.X, padx=10, pady=5)
        
        list_label = ttk.Label(self.model_list_frame, text="选择历史模型：")
        list_label.pack(anchor=tk.W)
        
        self.model_listbox = Listbox(self.model_list_frame, height=4, selectmode=SINGLE)
        self.model_listbox.pack(fill=tk.X, pady=5)
        
        # 刷新按钮
        btn_frame = ttk.Frame(self.model_list_frame)
        btn_frame.pack(fill=tk.X)
        
        ttk.Button(btn_frame, text="刷新模型列表", command=self.refresh_model_list).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="浏览历史模型", command=self.browse_historical_models).pack(side=tk.LEFT, padx=5)
        
        # 默认隐藏模型列表
        self.model_list_frame.pack_forget()
        
        # 开始处理按钮
        start_frame = ttk.Frame(self.main_frame)
        start_frame.pack(pady=20)
        
        self.start_button = ttk.Button(start_frame, text="🚀 开始处理", 
                                      command=self.start_processing,
                                      style="Accent.TButton")
        self.start_button.pack(pady=10)
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.main_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, padx=20, pady=10)
        
        # 状态显示区
        status_frame = ttk.LabelFrame(self.main_frame, text="处理状态")
        status_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.status_text = scrolledtext.ScrolledText(status_frame, height=15, wrap=tk.WORD)
        self.status_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.status_text.config(state=tk.DISABLED)
        
        # 底部按钮
        bottom_frame = ttk.Frame(self.main_frame)
        bottom_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(bottom_frame, text="退出程序", command=self.exit_program).pack(side=tk.RIGHT, padx=5)
    
    def create_columns_tab(self):
        """创建列选择页"""
        try:
            # 读取表头 - 修复：指定引擎
            try:
                df = pd.read_excel(HISTORY_EXCEL, nrows=0, dtype=str, engine='openpyxl')
            except Exception:
                try:
                    df = pd.read_excel(HISTORY_EXCEL, nrows=0, dtype=str, engine='xlrd')
                except Exception as e:
                    raise ValueError(f"无法读取历史数据：{str(e)}")
            
            all_cols = df.columns.tolist()
            
            if not all_cols:
                ttk.Label(self.columns_frame, text="历史数据无有效列！", font=("微软雅黑", 12)).pack(pady=50)
                return
        except Exception as e:
            ttk.Label(self.columns_frame, text=f"无法读取历史数据：{str(e)}", font=("微软雅黑", 12)).pack(pady=50)
            return
        
        # 加载历史配置
        saved_feat = [c for c in self.feat_cols if c in all_cols]
        saved_target = [c for c in self.target_cols if c in all_cols]
        saved_text = [c for c in self.text_cols if c in all_cols and c not in saved_target]
        
        # 变量定义
        self.feat_vars = {col: tk.BooleanVar(value=col in saved_feat) for col in all_cols}
        self.text_vars = {col: tk.BooleanVar(value=col in saved_text) for col in all_cols}
        self.target_vars = {col: tk.BooleanVar(value=col in saved_target) for col in all_cols}
        
        # 创建滚动区域
        canvas = tk.Canvas(self.columns_frame)
        scrollbar = ttk.Scrollbar(self.columns_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 1. 普通特征列
        feat_frame = ttk.LabelFrame(scrollable_frame, text="普通特征列（仅模型预测需选）")
        feat_frame.pack(fill=tk.X, padx=10, pady=10)
        
        feat_inner = ttk.Frame(feat_frame)
        feat_inner.pack(fill=tk.X, padx=10, pady=5)
        
        # 横板排列复选框（每行5列）
        current_row = None
        for i, col in enumerate(all_cols):
            if i % 5 == 0:
                current_row = ttk.Frame(feat_inner)
                current_row.pack(fill=tk.X, pady=2)
            
            cb = ttk.Checkbutton(current_row, text=col, variable=self.feat_vars[col])
            cb.pack(side=tk.LEFT, padx=5, pady=2)
        
        # 2. 文本特征列
        text_frame = ttk.LabelFrame(scrollable_frame, text="文本特征列（仅模型预测需选）")
        text_frame.pack(fill=tk.X, padx=10, pady=10)
        
        text_inner = ttk.Frame(text_frame)
        text_inner.pack(fill=tk.X, padx=10, pady=5)
        
        # 横板排列复选框（每行5列）
        current_row = None
        for i, col in enumerate(all_cols):
            if i % 5 == 0:
                current_row = ttk.Frame(text_inner)
                current_row.pack(fill=tk.X, pady=2)
            
            cb = ttk.Checkbutton(current_row, text=col, variable=self.text_vars[col])
            cb.pack(side=tk.LEFT, padx=5, pady=2)
        
        # 3. 目标列
        target_frame = ttk.LabelFrame(scrollable_frame, text="目标列（必须选择至少1个，可与查表目标列同列）")
        target_frame.pack(fill=tk.X, padx=10, pady=10)
        
        target_inner = ttk.Frame(target_frame)
        target_inner.pack(fill=tk.X, padx=10, pady=5)
        
        # 横板排列复选框（每行5列）
        current_row = None
        for i, col in enumerate(all_cols):
            if i % 5 == 0:
                current_row = ttk.Frame(target_inner)
                current_row.pack(fill=tk.X, pady=2)
            
            cb = ttk.Checkbutton(current_row, text=col, variable=self.target_vars[col])
            cb.pack(side=tk.LEFT, padx=5, pady=2)
        
        # 全选/取消按钮
        btn_frame = ttk.Frame(scrollable_frame)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(btn_frame, text="普通特征-全选", 
                  command=self.select_all_feat).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="普通特征-取消", 
                  command=self.deselect_all_feat).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(btn_frame, text="文本特征-全选", 
                  command=self.select_all_text).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="文本特征-取消", 
                  command=self.deselect_all_text).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(btn_frame, text="目标列-全选", 
                  command=self.select_all_target).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="目标列-取消", 
                  command=self.deselect_all_target).pack(side=tk.LEFT, padx=5)
        
        # 保存按钮
        save_frame = ttk.Frame(scrollable_frame)
        save_frame.pack(fill=tk.X, padx=10, pady=20)
        
        ttk.Button(save_frame, text="保存列配置", 
                  command=self.save_columns_config,
                  style="Accent.TButton").pack()
        
        # 打包滚动区域
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def create_mapping_tab(self):
        """创建映射配置页"""
        # 顶部说明
        info_frame = ttk.Frame(self.mapping_frame)
        info_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(info_frame, 
                 text="配置查表映射规则：选择映射源文件、键列（可多选）和值列（单选），然后添加到规则列表",
                 font=("微软雅黑", 10)).pack()
        
        # 主内容区域
        content_frame = ttk.Frame(self.mapping_frame)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 左半部分：列选择
        left_frame = ttk.Frame(content_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        
        # 文件选择 - 新增：显示当前选择的映射源文件
        file_frame = ttk.LabelFrame(left_frame, text="映射源文件")
        file_frame.pack(fill=tk.X, pady=5)
        
        self.mapping_file_var = tk.StringVar(value=HISTORY_EXCEL)
        file_entry = ttk.Entry(file_frame, textvariable=self.mapping_file_var, state='readonly')
        file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5, pady=5)
        
        ttk.Button(file_frame, text="浏览", 
                  command=self.browse_mapping_file).pack(side=tk.RIGHT, padx=5)
        
        # 当前文件显示
        current_file_frame = ttk.LabelFrame(left_frame, text="当前选择的映射源文件")
        current_file_frame.pack(fill=tk.X, pady=5)
        
        self.current_file_label = ttk.Label(current_file_frame, text=HISTORY_EXCEL, wraplength=300)
        self.current_file_label.pack(padx=5, pady=5)
        
        # 列列表
        cols_frame = ttk.LabelFrame(left_frame, text="文件列列表")
        cols_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # 修复：设置selectmode为EXTENDED，这样不按Ctrl也能多选
        self.columns_listbox = Listbox(cols_frame, selectmode="extended", height=15)
        self.columns_listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 加载列按钮
        ttk.Button(left_frame, text="加载文件列", 
                  command=self.load_file_columns).pack(pady=5)
        
        # 右半部分：映射配置
        right_frame = ttk.Frame(content_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5)
        
        # 键列选择
        key_frame = ttk.LabelFrame(right_frame, text="键列选择（按住Ctrl多选）")
        key_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # 修复：设置selectmode为EXTENDED
        self.key_listbox = Listbox(key_frame, selectmode="extended", height=8)
        self.key_listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 值列选择
        value_frame = ttk.LabelFrame(right_frame, text="值列选择（单选）")
        value_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.value_listbox = Listbox(value_frame, selectmode=SINGLE, height=8)
        self.value_listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 操作按钮
        btn_frame = ttk.Frame(right_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(btn_frame, text="添加选中列为键列", 
                  command=self.add_to_keys).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="添加选中列为值列", 
                  command=self.add_to_values).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="清空键列", 
                  command=lambda: self.key_listbox.delete(0, END)).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="清空值列", 
                  command=lambda: self.value_listbox.delete(0, END)).pack(side=tk.LEFT, padx=2)
        
        # 映射规则列表
        rules_frame = ttk.LabelFrame(self.mapping_frame, text="已配置的映射规则")
        rules_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 规则列表和操作按钮
        rules_btn_frame = ttk.Frame(rules_frame)
        rules_btn_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(rules_btn_frame, text="添加映射规则", 
                  command=self.add_mapping_rule).pack(side=tk.LEFT, padx=2)
        ttk.Button(rules_btn_frame, text="删除选中规则", 
                  command=self.delete_mapping_rule).pack(side=tk.LEFT, padx=2)
        ttk.Button(rules_btn_frame, text="清空所有规则", 
                  command=self.clear_mapping_rules).pack(side=tk.LEFT, padx=2)
        
        # 规则列表
        self.rules_listbox = Listbox(rules_frame, height=10)
        self.rules_listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 初始加载列
        self.load_file_columns()
    
    def create_params_tab(self):
        """创建参数配置页"""
        # 创建滚动区域
        canvas = tk.Canvas(self.params_frame)
        scrollbar = ttk.Scrollbar(self.params_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 随机森林参数
        rf_frame = ttk.LabelFrame(scrollable_frame, text="随机森林参数")
        rf_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # n_estimators
        n_est_frame = ttk.Frame(rf_frame)
        n_est_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(n_est_frame, text="决策树数量 (n_estimators):", width=25).pack(side=tk.LEFT)
        self.n_estimators_var = tk.StringVar(value="100")
        ttk.Entry(n_est_frame, textvariable=self.n_estimators_var, width=15).pack(side=tk.LEFT)
        ttk.Label(n_est_frame, text="默认：100，范围：50-500").pack(side=tk.LEFT, padx=10)
        
        # max_depth
        max_depth_frame = ttk.Frame(rf_frame)
        max_depth_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(max_depth_frame, text="树最大深度 (max_depth):", width=25).pack(side=tk.LEFT)
        self.max_depth_var = tk.StringVar(value="")
        ttk.Entry(max_depth_frame, textvariable=self.max_depth_var, width=15).pack(side=tk.LEFT)
        ttk.Label(max_depth_frame, text="留空=不限制，推荐：5-30").pack(side=tk.LEFT, padx=10)
        
        # min_samples_split
        min_split_frame = ttk.Frame(rf_frame)
        min_split_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(min_split_frame, text="节点分裂最小样本数:", width=25).pack(side=tk.LEFT)
        self.min_samples_split_var = tk.StringVar(value="")
        ttk.Entry(min_split_frame, textvariable=self.min_samples_split_var, width=15).pack(side=tk.LEFT)
        ttk.Label(min_split_frame, text="留空=2，范围：≥2").pack(side=tk.LEFT, padx=10)
        
        # min_samples_leaf
        min_leaf_frame = ttk.Frame(rf_frame)
        min_leaf_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(min_leaf_frame, text="叶子节点最小样本数:", width=25).pack(side=tk.LEFT)
        self.min_samples_leaf_var = tk.StringVar(value="")
        ttk.Entry(min_leaf_frame, textvariable=self.min_samples_leaf_var, width=15).pack(side=tk.LEFT)
        ttk.Label(min_leaf_frame, text="留空=1，范围：≥1").pack(side=tk.LEFT, padx=10)
        
        # random_state
        random_state_frame = ttk.Frame(rf_frame)
        random_state_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(random_state_frame, text="随机种子 (random_state):", width=25).pack(side=tk.LEFT)
        self.random_state_var = tk.StringVar(value="42")
        ttk.Entry(random_state_frame, textvariable=self.random_state_var, width=15).pack(side=tk.LEFT)
        ttk.Label(random_state_frame, text="默认：42，保证结果可复现").pack(side=tk.LEFT, padx=10)
        
        # TF-IDF参数
        tfidf_frame = ttk.LabelFrame(scrollable_frame, text="TF-IDF参数（文本特征提取）")
        tfidf_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # ngram_range
        ngram_frame = ttk.Frame(tfidf_frame)
        ngram_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(ngram_frame, text="n元语法范围 (ngram_range):", width=25).pack(side=tk.LEFT)
        
        ngram_inner = ttk.Frame(ngram_frame)
        ngram_inner.pack(side=tk.LEFT)
        
        self.ngram_min_var = tk.StringVar(value="1")
        ttk.Entry(ngram_inner, textvariable=self.ngram_min_var, width=5).pack(side=tk.LEFT)
        ttk.Label(ngram_inner, text="~").pack(side=tk.LEFT, padx=2)
        self.ngram_max_var = tk.StringVar(value="2")
        ttk.Entry(ngram_inner, textvariable=self.ngram_max_var, width=5).pack(side=tk.LEFT)
        
        ttk.Label(ngram_frame, text="默认：(1,2)").pack(side=tk.LEFT, padx=10)
        
        # max_features
        max_feat_frame = ttk.Frame(tfidf_frame)
        max_feat_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(max_feat_frame, text="最大特征数 (max_features):", width=25).pack(side=tk.LEFT)
        self.max_features_var = tk.StringVar(value="100")
        ttk.Entry(max_feat_frame, textvariable=self.max_features_var, width=15).pack(side=tk.LEFT)
        ttk.Label(max_feat_frame, text="默认：100，范围：>0").pack(side=tk.LEFT, padx=10)
        
        # stop_words
        stop_words_frame = ttk.Frame(tfidf_frame)
        stop_words_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(stop_words_frame, text="停用词 (stop_words):", width=25).pack(side=tk.LEFT)
        self.stop_words_var = tk.StringVar(value="english")
        ttk.Combobox(stop_words_frame, textvariable=self.stop_words_var,
                    values=["english", "None", "custom"], width=15).pack(side=tk.LEFT)
        ttk.Label(stop_words_frame, text="默认：english").pack(side=tk.LEFT, padx=10)
        
        # 置信度阈值
        threshold_frame = ttk.LabelFrame(scrollable_frame, text="置信度阈值")
        threshold_frame.pack(fill=tk.X, padx=10, pady=10)
        
        threshold_inner = ttk.Frame(threshold_frame)
        threshold_inner.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(threshold_inner, text="置信度阈值:", width=25).pack(side=tk.LEFT)
        self.confidence_threshold_var = tk.DoubleVar(value=CONFIDENCE_THRESHOLD)
        threshold_slider = ttk.Scale(threshold_inner, from_=0.5, to=1.0, 
                                    variable=self.confidence_threshold_var,
                                    orient=tk.HORIZONTAL, length=200)
        threshold_slider.pack(side=tk.LEFT, padx=10)
        
        threshold_label = ttk.Label(threshold_inner, text=f"{CONFIDENCE_THRESHOLD}")
        threshold_label.pack(side=tk.LEFT, padx=5)
        
        def update_threshold_label(val):
            threshold_label.config(text=f"{float(val):.2f}")
            self.filter.threshold = float(val)
        
        threshold_slider.config(command=update_threshold_label)
        
        # 按钮区域
        btn_frame = ttk.Frame(scrollable_frame)
        btn_frame.pack(fill=tk.X, padx=10, pady=20)
        
        ttk.Button(btn_frame, text="保存参数配置", 
                  command=self.save_params_config,
                  style="Accent.TButton").pack(pady=5)
        
        ttk.Button(btn_frame, text="重置为默认值", 
                  command=self.reset_params_to_default).pack(pady=5)
        
        # 打包滚动区域
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def create_log_tab(self):
        """创建日志管理页"""
        log_frame = ttk.Frame(self.log_frame)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 按钮区
        btn_frame = ttk.Frame(log_frame)
        btn_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(btn_frame, text="查看日志", command=self.logger.show_log_gui, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="打开日志文件", command=self.open_log_file, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="清空日志", command=self.clear_logs, width=15).pack(side=tk.LEFT, padx=5)
        
        # 日志显示区域
        text_frame = ttk.LabelFrame(log_frame, text="最近操作日志")
        text_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        self.log_text = scrolledtext.ScrolledText(text_frame, wrap=tk.WORD, font=("Consolas", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log_text.config(state=tk.DISABLED)
        
        # 刷新按钮
        ttk.Button(log_frame, text="刷新日志显示", command=self.update_log_display).pack(pady=5)
        
        # 初始显示日志
        self.update_log_display()
    
    def create_filter_tab(self):
        """创建低置信筛选页"""
        filter_frame = ttk.Frame(self.filter_frame)
        filter_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 设置区域
        settings_frame = ttk.LabelFrame(filter_frame, text="筛选设置")
        settings_frame.pack(fill=tk.X, pady=10)
        
        # 阈值设置
        threshold_frame = ttk.Frame(settings_frame)
        threshold_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(threshold_frame, text="置信度阈值：").pack(side=tk.LEFT)
        self.filter_threshold_var = tk.DoubleVar(value=CONFIDENCE_THRESHOLD)
        threshold_slider = ttk.Scale(threshold_frame, from_=0.5, to=1.0, 
                                    variable=self.filter_threshold_var,
                                    orient=tk.HORIZONTAL, length=200)
        threshold_slider.pack(side=tk.LEFT, padx=10)
        
        self.filter_threshold_label = ttk.Label(threshold_frame, text=f"{CONFIDENCE_THRESHOLD:.2f}")
        self.filter_threshold_label.pack(side=tk.LEFT, padx=5)
        
        def update_filter_threshold_label(val):
            self.filter_threshold_label.config(text=f"{float(val):.2f}")
            self.filter.threshold = float(val)
        
        threshold_slider.config(command=update_filter_threshold_label)
        
        # 文件选择
        file_frame = ttk.Frame(settings_frame)
        file_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(file_frame, text="输入文件：").pack(side=tk.LEFT)
        self.filter_input_var = tk.StringVar(value=NEW_EXCEL)
        ttk.Entry(file_frame, textvariable=self.filter_input_var, width=50).pack(side=tk.LEFT, padx=5)
        ttk.Button(file_frame, text="浏览", command=self.browse_filter_input).pack(side=tk.LEFT)
        
        # 操作按钮
        btn_frame = ttk.Frame(settings_frame)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(btn_frame, text="开始筛选", 
                  command=self.start_filtering,
                  style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        
        ttk.Button(btn_frame, text="打开结果文件", 
                  command=self.open_filter_result).pack(side=tk.LEFT, padx=5)
        
        # 结果显示区域
        result_frame = ttk.LabelFrame(filter_frame, text="筛选结果")
        result_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        self.filter_result_text = scrolledtext.ScrolledText(result_frame, wrap=tk.WORD, font=("Consolas", 9))
        self.filter_result_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.filter_result_text.config(state=tk.DISABLED)
    
    def create_incremental_tab(self):
        """创建增量学习页"""
        # 主框架
        main_frame = ttk.Frame(self.incremental_frame)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 1. 配置区域
        config_frame = ttk.LabelFrame(main_frame, text="增量学习配置")
        config_frame.pack(fill=tk.X, pady=5)
        
        # 自动增量学习开关
        switch_frame = ttk.Frame(config_frame)
        switch_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.auto_incremental_var = tk.BooleanVar(
            value=self.incremental_learner.config.get("auto_incremental_learning", True))
        
        ttk.Checkbutton(switch_frame, text="自动提取增量数据", 
                       variable=self.auto_incremental_var,
                       command=self.save_incremental_config).pack(side=tk.LEFT)
        
        # 保护已编辑数据开关
        protect_frame = ttk.Frame(config_frame)
        protect_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.protect_edited_var = tk.BooleanVar(
            value=self.incremental_learner.config.get("preserve_edited_data", True))
        
        ttk.Checkbutton(protect_frame, text="保护已编辑数据（手动提取时不覆盖已标记数据）", 
                       variable=self.protect_edited_var,
                       command=self.save_incremental_config).pack(side=tk.LEFT)
        
        # 阈值设置
        threshold_frame = ttk.Frame(config_frame)
        threshold_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(threshold_frame, text="置信度阈值：").pack(side=tk.LEFT)
        self.incremental_threshold_var = tk.DoubleVar(
            value=self.incremental_learner.config.get("confidence_threshold", CONFIDENCE_THRESHOLD))
        
        threshold_slider = ttk.Scale(threshold_frame, from_=0.5, to=1.0,
                                    variable=self.incremental_threshold_var,
                                    orient=tk.HORIZONTAL, length=200)
        threshold_slider.pack(side=tk.LEFT, padx=10)
        
        self.threshold_label = ttk.Label(threshold_frame, 
                                        text=f"{self.incremental_threshold_var.get():.2f}")
        self.threshold_label.pack(side=tk.LEFT)
        
        def update_threshold_label(val):
            self.threshold_label.config(text=f"{float(val):.2f}")
        
        threshold_slider.config(command=update_threshold_label)
        
        # 2. 操作按钮区域
        action_frame = ttk.LabelFrame(main_frame, text="增量学习操作")
        action_frame.pack(fill=tk.X, pady=5)
        
        btn_frame = ttk.Frame(action_frame)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(btn_frame, text="手动提取增量数据", 
                  command=self.manual_extract_incremental,
                  width=20).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(btn_frame, text="打开人工复核管理", 
                  command=self.open_human_review,
                  width=20).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(btn_frame, text="追加到历史数据", 
                  command=self.append_reviewed_to_history,
                  width=20).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(btn_frame, text="重新训练模型", 
                  command=self.retrain_with_incremental,
                  width=20).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(btn_frame, text="清空已复核数据", 
                  command=self.clear_reviewed_data,
                  width=20).pack(side=tk.LEFT, padx=5)
        
        # 3. 统计信息区域
        stats_frame = ttk.LabelFrame(main_frame, text="增量学习统计")
        stats_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.incremental_stats_text = scrolledtext.ScrolledText(
            stats_frame, wrap=tk.WORD, font=("微软雅黑", 10), height=15)
        self.incremental_stats_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.incremental_stats_text.config(state=tk.DISABLED)
        
        # 4. 文件操作区域
        file_frame = ttk.LabelFrame(main_frame, text="文件操作")
        file_frame.pack(fill=tk.X, pady=5)
        
        file_btn_frame = ttk.Frame(file_frame)
        file_btn_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(file_btn_frame, text="打开待复核文件", 
                  command=lambda: self.open_file(self.incremental_learner.pending_file),
                  width=15).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(file_btn_frame, text="打开已复核文件", 
                  command=lambda: self.open_file(self.incremental_learner.reviewed_file),
                  width=15).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(file_btn_frame, text="打开增量文件夹", 
                  command=lambda: self.open_folder(INCREMENTAL_DIR),
                  width=15).pack(side=tk.LEFT, padx=5)
        
        # 初始加载统计信息
        self.update_incremental_stats()
    
    # ==================== 功能方法 ====================
    
    def on_model_choice_changed(self):
        """模型选择变化事件"""
        if self.model_choice_var.get() == "use_existing":
            # 显示模型列表
            self.model_list_frame.pack(fill=tk.X, padx=10, pady=5)
            self.refresh_model_list()
        else:
            # 隐藏模型列表
            self.model_list_frame.pack_forget()
    
    def refresh_model_list(self):
        """刷新模型列表"""
        self.model_listbox.delete(0, END)
        
        models = self.version_manager.get_available_models()
        
        if not models:
            self.model_listbox.insert(END, "未找到可用的历史模型")
            self.model_listbox.config(state=tk.DISABLED)
            return
        
        self.model_listbox.config(state=tk.NORMAL)
        
        for model in models:
            version = model.get("version", "未知版本")
            timestamp = model.get("timestamp", "未知时间")
            display_text = f"{version} - {timestamp}"
            self.model_listbox.insert(END, display_text)
    
    def browse_historical_models(self):
        """浏览历史模型"""
        selection_window = ModelSelectionWindow(self.root, self.version_manager)
        self.root.wait_window(selection_window.window)
        
        selected_model = selection_window.get_selected_model()
        if selected_model:
            # 使用选中的模型
            try:
                encoder, tfidf_dict, count_dict, classifier, params, mapping_config = self.version_manager.load_model(selected_model["path"])
                self.current_model = (encoder, tfidf_dict, count_dict, classifier)
                self.model_params = params
                self.mapping_config = mapping_config
                
                # 更新模型列表显示
                self.refresh_model_list()
                
                # 更新状态
                self.update_status(f"✅ 已加载历史模型：{selected_model.get('version', '未知版本')}")
                messagebox.showinfo("成功", f"模型加载成功！\n版本：{selected_model.get('version', '未知')}\n时间：{selected_model.get('timestamp', '未知')}")
                
            except Exception as e:
                messagebox.showerror("加载失败", f"加载模型失败：{str(e)}")
    
    def check_files(self):
        """检查文件是否存在"""
        history_exists = os.path.exists(HISTORY_EXCEL)
        new_exists = os.path.exists(NEW_EXCEL)
        
        if history_exists:
            self.history_file_status.config(text="✅ 历史数据文件存在")
        else:
            self.history_file_status.config(text="❌ 历史数据文件不存在")
        
        if new_exists:
            self.new_file_status.config(text="✅ 新数据文件存在")
        else:
            self.new_file_status.config(text="❌ 新数据文件不存在")
        
        if history_exists and new_exists:
            self.update_status("文件检查完成，所有文件都存在")
        else:
            self.update_status("文件检查完成，部分文件缺失", warning=True)
    
    def select_all_feat(self):
        """普通特征-全选"""
        for var in self.feat_vars.values():
            var.set(True)
    
    def deselect_all_feat(self):
        """普通特征-取消全选"""
        for var in self.feat_vars.values():
            var.set(False)
    
    def select_all_text(self):
        """文本特征-全选"""
        for var in self.text_vars.values():
            var.set(True)
    
    def deselect_all_text(self):
        """文本特征-取消全选"""
        for var in self.text_vars.values():
            var.set(False)
    
    def select_all_target(self):
        """目标列-全选"""
        for var in self.target_vars.values():
            var.set(True)
    
    def deselect_all_target(self):
        """目标列-取消全选"""
        for var in self.target_vars.values():
            var.set(False)
    
    def save_columns_config(self):
        """保存列配置"""
        try:
            # 获取选中的列
            all_cols = list(self.feat_vars.keys())
            self.feat_cols = [col for col, var in self.feat_vars.items() if var.get()]
            self.text_cols = [col for col, var in self.text_vars.items() if var.get()]
            self.target_cols = [col for col, var in self.target_vars.items() if var.get()]
            
            if not self.target_cols:
                messagebox.showerror("错误", "至少选择1个目标列！")
                return
            
            # 保存配置
            save_config(self.feat_cols, self.target_cols, self.text_cols)
            
            # 更新状态
            self.update_status(f"列配置已保存：{len(self.feat_cols)}个特征列，{len(self.text_cols)}个文本列，{len(self.target_cols)}个目标列")
            messagebox.showinfo("成功", "列配置已保存！")
            
        except Exception as e:
            messagebox.showerror("保存失败", str(e))
    
    def browse_mapping_file(self):
        """浏览映射文件"""
        filename = filedialog.askopenfilename(
            title="选择映射源文件",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("All files", "*.*")],
            initialdir=DATA_DIR
        )
        if filename:
            self.mapping_file_var.set(filename)
            self.current_file_label.config(text=os.path.basename(filename))
            self.load_file_columns()
    
    def load_file_columns(self):
        """加载文件列"""
        try:
            file_path = self.mapping_file_var.get()
            if not file_path:
                messagebox.showwarning("提示", "请先选择文件")
                return
            
            # 修复：指定引擎
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path, nrows=0)
            else:
                try:
                    df = pd.read_excel(file_path, nrows=0, engine='openpyxl')
                except Exception:
                    try:
                        df = pd.read_excel(file_path, nrows=0, engine='xlrd')
                    except Exception as e:
                        raise ValueError(f"无法读取文件：{str(e)}")
            
            columns = df.columns.tolist()
            
            # 清空并更新列列表
            self.columns_listbox.delete(0, END)
            for col in columns:
                self.columns_listbox.insert(END, col)
            
            self.update_status(f"加载了 {len(columns)} 个列")
            
        except Exception as e:
            messagebox.showerror("加载失败", f"无法读取文件：{str(e)}")
    
    def add_to_keys(self):
        """添加选中列为键列"""
        selected = self.columns_listbox.curselection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要添加的列")
            return
        
        for index in selected:
            col = self.columns_listbox.get(index)
            # 检查是否已存在
            items = self.key_listbox.get(0, END)
            if col not in items:
                self.key_listbox.insert(END, col)
    
    def add_to_values(self):
        """添加选中列为值列"""
        selected = self.columns_listbox.curselection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要添加的列")
            return
        
        # 值列只能有一个，先清空
        self.value_listbox.delete(0, END)
        
        # 添加选中的列
        col = self.columns_listbox.get(selected[0])
        self.value_listbox.insert(END, col)
    
    def add_mapping_rule(self):
        """添加映射规则（现在包含来源文件信息）"""
        source_file = self.mapping_file_var.get()
        key_cols = list(self.key_listbox.get(0, END))
        value_cols = list(self.value_listbox.get(0, END))
        
        if not source_file:
            messagebox.showwarning("提示", "请先选择映射源文件")
            return
        
        if not key_cols:
            messagebox.showwarning("提示", "请选择至少一个键列")
            return
        
        if not value_cols:
            messagebox.showwarning("提示", "请选择一个值列")
            return
        
        value_col = value_cols[0]
        
        # 检查值列是否在键列中
        if value_col in key_cols:
            messagebox.showerror("错误", "值列不能同时作为键列")
            return
        
        # 添加到规则列表（显示格式：来源文件 → 键列 → 值列）
        source_name = os.path.basename(source_file)
        rule_str = f"来源：{source_name} | 键列：{', '.join(key_cols)} → 值列：{value_col}"
        self.rules_listbox.insert(END, rule_str)
        
        # 添加到配置（现在包含来源文件）
        self.mapping_config.append((source_file, key_cols.copy(), value_col))
        
        # 清空选择
        self.key_listbox.delete(0, END)
        self.value_listbox.delete(0, END)
        
        self.update_status(f"添加映射规则：{rule_str}")
    
    def delete_mapping_rule(self):
        """删除选中规则"""
        selected = self.rules_listbox.curselection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要删除的规则")
            return
        
        # 从后往前删除
        for index in reversed(selected):
            self.rules_listbox.delete(index)
            if index < len(self.mapping_config):
                self.mapping_config.pop(index)
        
        self.update_status(f"删除了 {len(selected)} 条规则")
    
    def clear_mapping_rules(self):
        """清空所有规则"""
        if messagebox.askyesno("确认", "确定要清空所有映射规则吗？"):
            self.rules_listbox.delete(0, END)
            self.mapping_config.clear()
            self.update_status("已清空所有映射规则")
    
    def save_params_config(self):
        """保存参数配置"""
        try:
            # 获取参数值
            rf_params = {
                "n_estimators": int(self.n_estimators_var.get()) if self.n_estimators_var.get().strip() else 100,
                "max_depth": int(self.max_depth_var.get()) if self.max_depth_var.get().strip() else None,
                "min_samples_split": int(self.min_samples_split_var.get()) if self.min_samples_split_var.get().strip() else 2,
                "min_samples_leaf": int(self.min_samples_leaf_var.get()) if self.min_samples_leaf_var.get().strip() else 1,
                "random_state": int(self.random_state_var.get()) if self.random_state_var.get().strip() else 42,
                "n_jobs": -1
            }
            
            tfidf_params = {
                "ngram_range": (
                    int(self.ngram_min_var.get()) if self.ngram_min_var.get().strip() else 1,
                    int(self.ngram_max_var.get()) if self.ngram_max_var.get().strip() else 2
                ),
                "max_features": int(self.max_features_var.get()) if self.max_features_var.get().strip() else 100,
                "stop_words": self.stop_words_var.get() if self.stop_words_var.get() != "None" else None
            }
            
            self.model_params = {
                "rf": rf_params,
                "tfidf": tfidf_params
            }
            
            self.update_status("参数配置已保存")
            messagebox.showinfo("成功", "参数配置已保存！")
            
        except Exception as e:
            messagebox.showerror("保存失败", f"参数格式错误：{str(e)}")
    
    def reset_params_to_default(self):
        """重置参数为默认值"""
        self.n_estimators_var.set("100")
        self.max_depth_var.set("")
        self.min_samples_split_var.set("")
        self.min_samples_leaf_var.set("")
        self.random_state_var.set("42")
        self.ngram_min_var.set("1")
        self.ngram_max_var.set("2")
        self.max_features_var.set("100")
        self.stop_words_var.set("english")
        self.confidence_threshold_var.set(CONFIDENCE_THRESHOLD)
        
        self.update_status("参数已重置为默认值")
    
    def start_processing(self):
        """开始处理"""
        progress_window = None
        try:
            # 检查配置
            if not self.target_cols:
                messagebox.showwarning("提示", "请先配置目标列（在'列选择'页）")
                self.notebook.select(1)  # 跳转到列选择页
                return
            
            if not self.model_params:
                messagebox.showwarning("提示", "请先配置模型参数（在'参数配置'页）")
                self.notebook.select(3)  # 跳转到参数配置页
                return
            
            # 验证文件存在
            if not os.path.exists(HISTORY_EXCEL):
                raise FileNotFoundError(f"历史数据文件不存在：{HISTORY_EXCEL}")
            
            if not os.path.exists(NEW_EXCEL):
                raise FileNotFoundError(f"新数据文件不存在：{NEW_EXCEL}")
            
            # 创建进度窗口
            progress_window = ProgressWindow(self.root, "数据处理中")
            
            def progress_callback(progress, status):
                """安全的进度回调函数"""
                try:
                    if progress_window and not progress_window.is_closed:
                        progress_window.update(progress, status)
                        self.progress_var.set(progress)
                        self.update_status(f"进度: {progress:.0f}% - {status}")
                except Exception:
                    pass  # 忽略进度更新错误
            
            # 开始处理
            progress_callback(5, "检查配置...")
            
            # 训练或加载模型
            if self.model_choice_var.get() == "train_new":
                progress_callback(10, "开始训练新模型...")
                encoder, tfidf_dict, count_dict, classifier, train_count = train_model(
                    self.feat_cols, self.text_cols, self.target_cols, 
                    self.model_params, HISTORY_EXCEL
                )
                self.current_model = (encoder, tfidf_dict, count_dict, classifier)
                
                # 备份模型和配置
                progress_callback(25, "备份模型和配置...")
                model_components = self.current_model
                version = self.version_manager.backup_model_and_mapping(
                    model_components, self.mapping_config, self.model_params
                )
                progress_callback(30, f"模型备份完成（版本：{version}）")
                
            else:
                # 使用历史模型
                progress_callback(10, "加载历史模型...")
                
                # 检查是否有选中的模型
                if not self.current_model:
                    # 尝试加载当前模型
                    try:
                        model_path = os.path.join(MODEL_DIR, "current_model.pkl")
                        if os.path.exists(model_path):
                            encoder, tfidf_dict, count_dict, classifier, params, mapping_config = self.version_manager.load_model(model_path)
                            self.current_model = (encoder, tfidf_dict, count_dict, classifier)
                            if params:
                                self.model_params = params
                            if mapping_config:
                                self.mapping_config = mapping_config
                            train_df = pd.read_excel(HISTORY_EXCEL, dtype=str)
                            train_count = len(train_df)
                            progress_callback(20, "历史模型加载成功")
                        else:
                            messagebox.showwarning("警告", "未找到历史模型，将训练新模型")
                            encoder, tfidf_dict, count_dict, classifier, train_count = train_model(
                                self.feat_cols, self.text_cols, self.target_cols, 
                                self.model_params, HISTORY_EXCEL
                            )
                            self.current_model = (encoder, tfidf_dict, count_dict, classifier)
                    except Exception as e:
                        messagebox.showwarning("警告", f"加载历史模型失败：{str(e)}，将训练新模型")
                        encoder, tfidf_dict, count_dict, classifier, train_count = train_model(
                            self.feat_cols, self.text_cols, self.target_cols, 
                            self.model_params, HISTORY_EXCEL
                        )
                        self.current_model = (encoder, tfidf_dict, count_dict, classifier)
                else:
                    # 使用已加载的模型
                    train_df = pd.read_excel(HISTORY_EXCEL, dtype=str)
                    train_count = len(train_df)
                    progress_callback(20, "使用已加载的模型")
            
            # 执行预测
            progress_callback(30, "开始预测...")
            encoder, tfidf_dict, count_dict, classifier = self.current_model
            
            result_df, predict_count = predict_data(
                encoder, tfidf_dict, count_dict, classifier,
                self.feat_cols, self.text_cols, self.target_cols,
                self.mapping_config, self.model_params, NEW_EXCEL,
                progress_callback
            )
            
            # 保存结果
            progress_callback(95, "保存结果到Excel...")
            write_results_to_excel(result_df, NEW_EXCEL)
            
            # 计算统计信息
            if "置信度" in result_df.columns:
                confidence_stats = {
                    "mean": float(pd.to_numeric(result_df["置信度"], errors='coerce').mean()),
                    "min": float(pd.to_numeric(result_df["置信度"], errors='coerce').min()),
                    "low_confidence_count": int(len(result_df[pd.to_numeric(result_df["置信度"], errors='coerce') < CONFIDENCE_THRESHOLD]))
                }
            else:
                confidence_stats = {"mean": 0.0, "min": 0.0, "low_confidence_count": 0}
            
            # 写入日志
            progress_callback(98, "写入操作日志...")
            self.logger.write_log(
                self.feat_cols, self.target_cols, train_count, predict_count,
                confidence_stats, self.mapping_config
            )
            
            # 完成
            if progress_window:
                progress_window.close()
            progress_callback(100, "处理完成！")
            
            # 显示结果
            self.update_status("✅ 处理完成！", success=True)
            
            result_message = (
                f"✅ 数据处理完成！\n\n"
                f"📊 处理结果：\n"
                f"- 训练数据：{train_count} 行\n"
                f"- 预测数据：{predict_count} 行\n"
            )
            
            if "置信度" in result_df.columns:
                result_message += (
                    f"- 平均置信度：{confidence_stats['mean']:.4f}\n"
                    f"- 最低置信度：{confidence_stats['min']:.4f}\n"
                    f"- 低置信度数据：{confidence_stats['low_confidence_count']} 条\n"
                )
            
            result_message += f"\n💾 结果已保存到：\n{NEW_EXCEL}"
            
            messagebox.showinfo("完成", result_message)
            
            # 增量学习流程
            if self.auto_incremental_var.get():
                try:
                    self.update_status("开始增量学习流程...")
                    
                    # 创建数据对比器
                    self.data_comparator = DataComparator(HISTORY_EXCEL, NEW_EXCEL)
                    
                    # 对比特征
                    feature_comparison = self.data_comparator.compare_features(
                        self.feat_cols, self.text_cols)
                    
                    # 对比结果
                    result_comparison = self.data_comparator.compare_results(
                        self.target_cols, result_df)
                    
                    # 提取增量数据 - 修复：避免重复提取已追加的数据
                    incremental_df = self.data_comparator.extract_incremental_data(
                        result_df, self.feat_cols + self.text_cols, self.target_cols,
                        confidence_threshold=self.incremental_threshold_var.get()
                    )
                    
                    if not incremental_df.empty:
                        # 过滤已追加的数据（通过配置中的appended_records）
                        config = self.incremental_learner.config
                        appended_records = config.get("appended_records", [])
                        
                        # 如果配置中有已追加记录，过滤掉这些数据
                        if appended_records:
                            # 计算每个数据的唯一标识
                            filtered_data = []
                            for idx, row in incremental_df.iterrows():
                                data_id = self.incremental_learner._create_data_id(
                                    row, [col for col in incremental_df.columns if not col.startswith('_')]
                                )
                                if data_id not in appended_records:
                                    filtered_data.append(row)
                            
                            if filtered_data:
                                incremental_df = pd.DataFrame(filtered_data)
                                self.update_status(f"过滤掉 {len(incremental_df) - len(filtered_data)} 条已追加数据")
                            else:
                                incremental_df = pd.DataFrame()
                                self.update_status("所有增量数据都已被追加过")
                        
                        # 保存增量数据
                        if not incremental_df.empty:
                            success, stats = self.incremental_learner.save_incremental_data(
                                incremental_df, "自动提取", preserve_edited=self.protect_edited_var.get()
                            )
                            
                            if success:
                                self.update_incremental_stats()
                                self.update_status(f"✅ 已提取 {len(incremental_df)} 条增量数据待复核")
                                
                                # 显示增量学习结果
                                incremental_message = (
                                    f"📊 增量学习结果：\n"
                                    f"- 提取增量数据：{len(incremental_df)} 条\n"
                                    f"- 总待复核数据：{stats['总待复核记录数']} 条\n"
                                    f"- 保护已编辑数据：{'是' if self.protect_edited_var.get() else '否'}\n"
                                )
                                
                                if feature_comparison:
                                    inc_features = feature_comparison["summary"]["total_differences"]
                                    if inc_features > 0:
                                        incremental_message += f"- 新特征值：{inc_features} 列\n"
                                
                                if result_comparison:
                                    total_new_cats = sum(
                                        result_comparison["summary"]["new_categories"].values())
                                    if total_new_cats > 0:
                                        incremental_message += f"- 新预测类别：{total_new_cats} 个\n"
                                
                                self.update_status(incremental_message)
                                
                                # 询问是否打开复核界面
                                if messagebox.askyesno("增量数据", 
                                    f"已提取 {len(incremental_df)} 条增量数据需要人工复核。\n"
                                    f"是否现在打开人工复核管理界面？"):
                                    self.open_human_review()
                        
                        else:
                            self.update_status("✅ 无新增增量数据需要复核")
                    
                    else:
                        self.update_status("✅ 未发现需要复核的增量数据")
                        
                except Exception as e:
                    self.update_status(f"⚠️ 增量学习失败：{str(e)}", warning=True)
                    # 增量学习失败不影响主流程
            
        except Exception as e:
            self.update_status(f"❌ 处理失败：{str(e)}", error=True)
            messagebox.showerror("处理失败", f"错误详情：{str(e)}\n\n请检查配置和数据文件。")
            if progress_window:
                progress_window.close()
    
    def save_incremental_config(self):
        """保存增量学习配置"""
        self.incremental_learner.config["auto_incremental_learning"] = self.auto_incremental_var.get()
        self.incremental_learner.config["confidence_threshold"] = self.incremental_threshold_var.get()
        self.incremental_learner.config["preserve_edited_data"] = self.protect_edited_var.get()
        self.incremental_learner.save_config()
        self.update_status("增量学习配置已保存")
    
    def manual_extract_incremental(self):
        """手动提取增量数据"""
        try:
            # 检查是否有新数据结果
            if not os.path.exists(NEW_EXCEL):
                messagebox.showwarning("提示", "请先运行主处理生成新数据结果")
                return
            
            # 读取结果数据
            result_df = pd.read_excel(NEW_EXCEL, dtype=str)
            
            # 创建数据对比器
            self.data_comparator = DataComparator(HISTORY_EXCEL, NEW_EXCEL)
            
            # 提取增量数据
            incremental_df = self.data_comparator.extract_incremental_data(
                result_df, self.feat_cols + self.text_cols, self.target_cols,
                confidence_threshold=self.incremental_threshold_var.get()
            )
            
            if incremental_df.empty:
                messagebox.showinfo("结果", "未发现需要复核的增量数据")
                return
            
            # 保存增量数据（根据配置决定是否保护已编辑数据）
            preserve_edited = self.protect_edited_var.get()
            success, stats = self.incremental_learner.save_incremental_data(
                incremental_df, "手动提取", preserve_edited=preserve_edited
            )
            
            if success:
                self.update_incremental_stats()
                messagebox.showinfo("成功", 
                    f"✅ 提取完成！\n"
                    f"新增 {stats['新增记录数']} 条增量数据\n"
                    f"总待复核记录：{stats['总待复核记录数']} 条\n"
                    f"保护已编辑数据：{'是' if preserve_edited else '否'}")
            else:
                messagebox.showerror("失败", stats)
                
        except Exception as e:
            messagebox.showerror("提取失败", f"提取增量数据失败：{str(e)}")
    
    def open_human_review(self):
        """打开人工复核管理窗口"""
        HumanReviewWindow(self.root, self.incremental_learner)
    
    def append_reviewed_to_history(self):
        """将已复核数据追加到历史数据"""
        if messagebox.askyesno("确认", 
            "确定要将已复核数据追加到历史数据吗？\n"
            "注意：追加时会自动移除置信度列，避免污染历史数据。\n"
            "追加后建议重新训练模型。"):
            
            success, message = self.incremental_learner.append_to_history(
                HISTORY_EXCEL, 
                remove_confidence_column=True,  # 移除置信度列
                reviewer_name="用户复核"
            )
            
            if success:
                self.update_incremental_stats()
                messagebox.showinfo("成功", message)
                
                # 询问是否重新训练模型
                if messagebox.askyesno("重新训练", "是否要重新训练模型以包含新增数据？"):
                    self.retrain_with_incremental()
            else:
                messagebox.showerror("失败", message)
    
    def clear_reviewed_data(self):
        """清空已复核数据"""
        if messagebox.askyesno("确认", "确定要清空所有已复核数据吗？此操作不可恢复！"):
            success, message = self.incremental_learner.clear_reviewed_data()
            if success:
                self.update_incremental_stats()
                messagebox.showinfo("成功", message)
            else:
                messagebox.showerror("失败", message)
    
    def retrain_with_incremental(self):
        """使用增量数据重新训练模型"""
        try:
            # 检查历史数据是否已更新
            if not os.path.exists(HISTORY_EXCEL):
                messagebox.showerror("错误", "历史数据文件不存在")
                return
            
            # 更新状态
            self.update_status("开始使用增量数据重新训练模型...")
            
            # 训练新模型
            encoder, tfidf_dict, count_dict, classifier, train_count = train_model(
                self.feat_cols, self.text_cols, self.target_cols, 
                self.model_params, HISTORY_EXCEL
            )
            
            # 更新当前模型
            self.current_model = (encoder, tfidf_dict, count_dict, classifier)
            
            # 备份模型和配置
            model_components = self.current_model
            version = self.version_manager.backup_model_and_mapping(
                model_components, self.mapping_config, self.model_params
            )
            
            self.update_status(f"✅ 模型重新训练完成（版本：{version}，训练数据：{train_count}行）")
            messagebox.showinfo("成功", 
                f"模型重新训练完成！\n"
                f"版本：{version}\n"
                f"训练数据：{train_count} 行\n"
                f"已包含所有增量数据")
            
        except Exception as e:
            self.update_status(f"❌ 重新训练失败：{str(e)}", error=True)
            messagebox.showerror("重新训练失败", str(e))
    
    def update_incremental_stats(self):
        """更新增量学习统计信息"""
        try:
            config = self.incremental_learner.config
            pending_df = self.incremental_learner.load_pending_data()
            reviewed_df = self.incremental_learner.load_reviewed_data()
            
            stats_text = f"📊 增量学习统计信息\n"
            stats_text += "="*60 + "\n"
            stats_text += f"自动提取状态：{'✅ 开启' if config.get('auto_incremental_learning', True) else '❌ 关闭'}\n"
            stats_text += f"置信度阈值：{config.get('confidence_threshold', CONFIDENCE_THRESHOLD):.2f}\n"
            stats_text += f"累计复核数据：{config.get('total_reviewed', 0):,} 条\n"
            stats_text += f"累计追加数据：{config.get('total_added', 0):,} 条\n"
            stats_text += f"最后复核日期：{config.get('last_review_date', '从未复核')}\n"
            stats_text += f"最后追加日期：{config.get('last_append_date', '从未追加')}\n"
            stats_text += f"保护已编辑数据：{'✅ 是' if config.get('preserve_edited_data', True) else '❌ 否'}\n"
            stats_text += "="*60 + "\n\n"
            
            stats_text += f"📋 待复核数据：{len(pending_df):,} 条\n"
            if not pending_df.empty:
                if "_复核原因" in pending_df.columns:
                    reason_stats = pending_df["_复核原因"].value_counts().to_dict()
                    for reason, count in reason_stats.items():
                        stats_text += f"  • {reason}: {count:,} 条\n"
                
                if "_复核状态" in pending_df.columns:
                    reviewed_count = pending_df["_复核状态"].notna().sum()
                    stats_text += f"  已标记数：{reviewed_count:,} 条\n"
                    stats_text += f"  待处理数：{len(pending_df) - reviewed_count:,} 条\n"
            
            stats_text += f"\n✅ 已复核数据：{len(reviewed_df):,} 条\n"
            if not reviewed_df.empty:
                if "_复核时间" in reviewed_df.columns:
                    reviewed_df["_复核日期"] = reviewed_df["_复核时间"].str[:10]
                    date_stats = reviewed_df["_复核日期"].value_counts().head(5).to_dict()
                    for date, count in date_stats.items():
                        stats_text += f"  • {date}: {count:,} 条\n"
            
            # 历史数据统计
            try:
                if os.path.exists(HISTORY_EXCEL):
                    history_df = pd.read_excel(HISTORY_EXCEL, dtype=str)
                    stats_text += f"\n📚 历史数据总计：{len(history_df):,} 条\n"
                    
                    # 按目标列统计
                    for target_col in self.target_cols:
                        if target_col in history_df.columns:
                            cat_count = history_df[target_col].nunique()
                            stats_text += f"  • {target_col}: {cat_count} 个类别\n"
            except:
                pass
            
            self.incremental_stats_text.config(state=tk.NORMAL)
            self.incremental_stats_text.delete(1.0, tk.END)
            self.incremental_stats_text.insert(tk.END, stats_text)
            self.incremental_stats_text.config(state=tk.DISABLED)
            
        except Exception as e:
            print(f"更新增量统计失败：{str(e)}")
    
    def open_file(self, filepath):
        """打开文件"""
        try:
            if os.path.exists(filepath):
                if os.name == 'nt':
                    os.startfile(filepath)
                else:
                    import subprocess
                    subprocess.run(['open' if sys.platform == 'darwin' else 'xdg-open', filepath])
            else:
                messagebox.showwarning("提示", "文件不存在")
        except Exception as e:
            messagebox.showerror("打开失败", f"无法打开文件：{str(e)}")
    
    def open_folder(self, folderpath):
        """打开文件夹"""
        try:
            if os.path.exists(folderpath):
                if os.name == 'nt':
                    os.startfile(folderpath)
                else:
                    import subprocess
                    subprocess.run(['open' if sys.platform == 'darwin' else 'xdg-open', folderpath])
            else:
                messagebox.showwarning("提示", "文件夹不存在")
        except Exception as e:
            messagebox.showerror("打开失败", f"无法打开文件夹：{str(e)}")
    
    def start_filtering(self):
        """开始筛选低置信数据"""
        try:
            input_path = self.filter_input_var.get()
            self.filter.threshold = self.filter_threshold_var.get()
            
            success, stats = self.filter.filter_low_confidence(input_path)
            
            if success:
                result_text = (f"✅ 筛选完成！\n\n"
                              f"📊 筛选结果：\n"
                              f"- 共筛选出 {stats['count']} 条低置信数据\n"
                              f"- 置信度范围：{stats['min_confidence']:.4f} ~ {stats['max_confidence']:.4f}\n"
                              f"- 平均置信度：{stats['avg_confidence']:.4f}\n\n"
                              f"📁 文件位置：\n{stats['output_path']}")
                
                self.filter_result_text.config(state=tk.NORMAL)
                self.filter_result_text.delete(1.0, tk.END)
                self.filter_result_text.insert(tk.END, result_text)
                self.filter_result_text.config(state=tk.DISABLED)
                
                self.update_status(f"筛选完成：{stats['count']} 条低置信数据")
                messagebox.showinfo("筛选成功", f"共筛选出 {stats['count']} 条低置信数据")
                
            elif "error" in stats:
                self.update_status(f"筛选失败：{stats['error']}", error=True)
                messagebox.showerror("筛选失败", stats['error'])
            else:
                result_text = "✅ 未找到低置信数据（所有数据置信度均达标）"
                self.filter_result_text.config(state=tk.NORMAL)
                self.filter_result_text.delete(1.0, tk.END)
                self.filter_result_text.insert(tk.END, result_text)
                self.filter_result_text.config(state=tk.DISABLED)
                
                self.update_status("未找到低置信数据")
                messagebox.showinfo("筛选结果", "未找到低置信数据")
                
        except Exception as e:
            self.update_status(f"筛选失败：{str(e)}", error=True)
            messagebox.showerror("筛选失败", str(e))
    
    def browse_filter_input(self):
        """浏览筛选输入文件"""
        filename = filedialog.askopenfilename(
            title="选择输入文件",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.filter_input_var.set(filename)
    
    def open_filter_result(self):
        """打开筛选结果文件"""
        try:
            if os.path.exists(LOW_CONF_OUTPUT):
                if os.name == 'nt':
                    os.startfile(LOW_CONF_OUTPUT)
                else:
                    import subprocess
                    subprocess.run(['open' if sys.platform == 'darwin' else 'xdg-open', LOW_CONF_OUTPUT])
            else:
                messagebox.showwarning("提示", "结果文件不存在，请先进行筛选")
        except Exception as e:
            messagebox.showerror("打开失败", f"无法打开文件：{str(e)}")
    
    def update_log_display(self):
        """更新日志显示"""
        try:
            with open(LOG_JSON_PATH, "r", encoding="utf-8") as f:
                logs = json.load(f)
            
            self.log_text.config(state=tk.NORMAL)
            self.log_text.delete(1.0, tk.END)
            
            if logs:
                # 显示最近10条日志
                recent_logs = logs[-10:] if len(logs) >= 10 else logs
                
                for log in reversed(recent_logs):  # 最新的在前面
                    self.log_text.insert(tk.END, f"📅 {log['timestamp']}\n")
                    self.log_text.insert(tk.END, f"   特征列：{', '.join(log['feature_columns'])}\n")
                    self.log_text.insert(tk.END, f"   目标列：{', '.join(log['target_columns'])}\n")
                    self.log_text.insert(tk.END, f"   训练数据：{log['train_data_rows']} 行\n")
                    self.log_text.insert(tk.END, f"   预测数据：{log['predict_data_rows']} 行\n")
                    
                    conf = log['confidence_statistics']
                    if isinstance(conf, dict):
                        self.log_text.insert(tk.END, f"   平均置信度：{conf.get('mean', 0.0):.4f}\n")
                        self.log_text.insert(tk.END, f"   低置信度数据：{conf.get('low_confidence_count', 0)} 条\n")
                    
                    self.log_text.insert(tk.END, "-"*60 + "\n\n")
            else:
                self.log_text.insert(tk.END, "暂无操作记录")
            
            self.log_text.config(state=tk.DISABLED)
            
        except Exception as e:
            self.log_text.config(state=tk.NORMAL)
            self.log_text.delete(1.0, tk.END)
            self.log_text.insert(tk.END, f"读取日志失败：{str(e)}")
            self.log_text.config(state=tk.DISABLED)
    
    def open_log_file(self):
        """打开日志文件"""
        try:
            # 修复：使用正确的方式打开文本文件，确保编码正确
            if os.name == 'nt':
                # Windows系统使用记事本打开
                os.system(f'notepad "{LOG_TXT_PATH}"')
            else:
                # Mac/Linux系统使用默认文本编辑器
                import subprocess
                subprocess.run(['open' if sys.platform == 'darwin' else 'xdg-open', LOG_TXT_PATH])
        except Exception as e:
            messagebox.showerror("打开失败", f"无法打开日志文件：{str(e)}")
    
    def clear_logs(self):
        """清空日志"""
        if messagebox.askyesno("确认", "确定要清空所有日志吗？此操作不可恢复！"):
            try:
                with open(LOG_JSON_PATH, "w", encoding="utf-8") as f:
                    json.dump([], f, ensure_ascii=False, indent=4)
                
                with open(LOG_TXT_PATH, "w", encoding="utf-8-sig") as f:
                    f.write("智能标签填充操作日志\n")
                    f.write("="*80 + "\n")
                
                self.update_log_display()
                self.update_status("✅ 日志已清空", success=True)
                messagebox.showinfo("成功", "日志已清空")
            except Exception as e:
                messagebox.showerror("清空失败", str(e))
    
    def update_status(self, message, success=False, warning=False, error=False):
        """更新状态"""
        self.status_text.config(state=tk.NORMAL)
        self.status_text.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
        
        # 添加样式
        if error:
            self.status_text.tag_add("error", "end-1c linestart", "end-1c lineend")
            self.status_text.tag_config("error", foreground="red")
        elif warning:
            self.status_text.tag_add("warning", "end-1c linestart", "end-1c lineend")
            self.status_text.tag_config("warning", foreground="orange")
        elif success:
            self.status_text.tag_add("success", "end-1c linestart", "end-1c lineend")
            self.status_text.tag_config("success", foreground="green")
        
        self.status_text.see(tk.END)
        self.status_text.config(state=tk.DISABLED)
        
        # 更新状态栏
        if hasattr(self, 'status_bar') and self.status_bar:
            self.status_bar.config(text=message)
    
    def on_closing(self):
        """窗口关闭事件"""
        if self.is_exiting:
            return
        
        self.is_exiting = True
        if messagebox.askyesno("退出", "确定要退出程序吗？"):
            self.root.quit()
            self.root.destroy()
            sys.exit(0)
        else:
            self.is_exiting = False
    
    def exit_program(self):
        """退出程序"""
        self.on_closing()
    
    def run(self):
        """运行程序"""
        try:
            # 设置样式
            style = ttk.Style()
            style.configure("Accent.TButton", font=("微软雅黑", 10, "bold"), padding=10)
            
            # 初始检查
            self.check_files()
            
            self.root.mainloop()
        except KeyboardInterrupt:
            self.root.quit()
        except Exception as e:
            messagebox.showerror("程序错误", f"程序运行出错：{str(e)}\n\n{traceback.format_exc()}")
            self.root.quit()

# ==================== 主程序入口 ====================
if __name__ == "__main__":
    # 检查必要文件
    if not os.path.exists(HISTORY_EXCEL):
        print(f"⚠️  警告：历史数据文件不存在，请创建：{HISTORY_EXCEL}")
    
    if not os.path.exists(NEW_EXCEL):
        print(f"⚠️  警告：新数据文件不存在，请创建：{NEW_EXCEL}")
    
    app = SmartLabelToolkit()
    app.run()